import pandas as pd
import tushare as ts
import numpy as np
import time
import sqlite3
from datetime import datetime
import os
from loguru import logger
import argparse
from tqdm import tqdm
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv
import config

class TokenManager:
    """管理多个Tushare token的切换和重试逻辑"""
    
    def __init__(self, tokens):
        self.tokens = tokens
        self.current_token_index = 0
        self.token_retry_count = {}
        self.total_requests = 0
        self.successful_requests = 0
        self.max_retries_per_token = 3
        self.token_switch_delay = 60  # 切换token后等待时间（秒）
        
        if not tokens:
            raise ValueError("至少需要提供一个Tushare token")
            
        logger.info(f"🔧 初始化Token管理器，共有 {len(tokens)} 个token可用")
        self._switch_token()
    
    def _switch_token(self):
        """切换到当前token"""
        if len(self.tokens) == 1:
            current_token = self.tokens[0]
        else:
            current_token = self.tokens[self.current_token_index]
            
        logger.info(f"🔄 切换到Token {self.current_token_index + 1}/{len(self.tokens)}")
        ts.set_token(current_token)
        self.pro = ts.pro_api()
        
        # 重置当前token的重试次数
        self.token_retry_count[self.current_token_index] = 0
    
    def _next_token(self):
        """切换到下一个可用token"""
        if len(self.tokens) <= 1:
            return False
            
        # 尝试切换到下一个token
        original_index = self.current_token_index
        while True:
            self.current_token_index = (self.current_token_index + 1) % len(self.tokens)
            
            # 如果回到原始token，说明所有token都试过了
            if self.current_token_index == original_index:
                return False
                
            # 检查这个token是否还有重试机会
            retry_count = self.token_retry_count.get(self.current_token_index, 0)
            if retry_count < self.max_retries_per_token:
                self._switch_token()
                time.sleep(self.token_switch_delay)  # 切换后等待
                return True
        
        return False
    
    def make_request(self, request_func, *args, **kwargs):
        """执行API请求，包含重试和token切换逻辑"""
        self.total_requests += 1
        
        while True:
            try:
                # 记录当前token的重试次数
                current_retry = self.token_retry_count.get(self.current_token_index, 0)
                
                if current_retry > 0:
                    logger.warning(f"⚠️  Token {self.current_token_index + 1} 重试第 {current_retry} 次")
                
                # 执行请求
                result = request_func(self.pro, *args, **kwargs)
                
                # 请求成功
                self.successful_requests += 1
                self.token_retry_count[self.current_token_index] = 0  # 重置重试次数
                return result
                
            except Exception as e:
                error_msg = str(e)
                self.token_retry_count[self.current_token_index] = current_retry + 1
                
                logger.error(f"❌ API请求失败 (Token {self.current_token_index + 1}, 重试 {current_retry + 1}): {error_msg}")
                
                # 检查是否是API限制错误
                if any(keyword in error_msg.lower() for keyword in ['limit', '限制', 'timeout', '超时', 'rate']):
                    logger.warning("🚦 检测到API限制，尝试切换token...")
                    
                    # 尝试切换token
                    if self._next_token():
                        logger.info(f"✅ 已切换到Token {self.current_token_index + 1}")
                        continue
                    else:
                        logger.warning("⚠️  所有token都已达到重试限制，等待后重置...")
                        time.sleep(self.token_switch_delay * 2)  # 等待更长时间
                        # 重置所有token的重试次数
                        self.token_retry_count = {}
                        self.current_token_index = 0
                        self._switch_token()
                        continue
                
                # 检查当前token是否还有重试机会
                if current_retry < self.max_retries_per_token:
                    wait_time = 2 ** current_retry  # 指数退避
                    logger.info(f"⏳ 等待 {wait_time} 秒后重试...")
                    time.sleep(wait_time)
                    continue
                else:
                    # 当前token重试次数已达上限，尝试切换
                    if self._next_token():
                        logger.info(f"✅ Token重试次数达上限，已切换到Token {self.current_token_index + 1}")
                        continue
                    else:
                        # 所有token都试过了，抛出异常
                        raise Exception(f"所有token都无法完成请求: {error_msg}")
    
    def get_stats(self):
        """获取请求统计信息"""
        success_rate = (self.successful_requests / self.total_requests * 100) if self.total_requests > 0 else 0
        return {
            'total_requests': self.total_requests,
            'successful_requests': self.successful_requests,
            'success_rate': f"{success_rate:.1f}%",
            'current_token': self.current_token_index + 1,
            'total_tokens': len(self.tokens)
        }

def check_environment():
    """检查环境变量是否正确设置"""
    load_dotenv()
    tushare_token = os.getenv('TUSHARE_TOKEN')
    if not tushare_token:
        raise ValueError("请在 .env 文件中设置 TUSHARE_TOKEN")
    return tushare_token

class StockDataCollector:
    def __init__(self, tokens, cache_dir='cache', batch_size=50, use_delay=True):
        # 初始化Token管理器
        self.token_manager = TokenManager(tokens)
        logger.info("Tushare API 初始化成功")
        
        # 创建缓存目录
        self.cache_dir = cache_dir
        self.batch_size = batch_size
        self.use_delay = use_delay  # 是否使用延时
        os.makedirs(cache_dir, exist_ok=True)
        
    def _get_batch_cache_path(self, batch_index):
        """获取批次缓存文件路径"""
        return os.path.join(self.cache_dir, f"batch_{batch_index}.json")
        
    def _save_batch_to_cache(self, batch_data, batch_index):
        """保存批次数据到缓存"""
        if not batch_data:
            return
            
        cache_path = self._get_batch_cache_path(batch_index)
        try:
            with open(cache_path, 'w', encoding='utf-8') as f:
                json.dump(batch_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"保存批次缓存失败 {cache_path}: {e}")
            
    def _load_batch_from_cache(self, batch_index):
        """从缓存加载批次数据"""
        cache_path = self._get_batch_cache_path(batch_index)
        if not os.path.exists(cache_path):
            return None
            
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"读取批次缓存失败 {cache_path}: {e}")
            return None

    def get_all_stocks(self):
        """获取所有A股上市公司列表"""
        try:
            # 从API获取数据
            stocks = self.token_manager.make_request(lambda pro: pro.stock_basic(exchange='', list_status='L'))
            return stocks[['ts_code', 'name', 'industry']]
        except Exception as e:
            logger.error(f"获取股票列表失败: {e}")
            return None

    def get_annual_data(self, stock_code, start_year, end_year):
        """获取单个股票的年度数据"""
        try:
            # 构建年份范围（排除当前年份，只获取完整年份的数据）
            current_year = datetime.now().year
            actual_end_year = min(end_year, current_year - 1)  # 不包含当前年份
            years = range(start_year, actual_end_year + 1)
            
            # 预筛选：检查最近3年是否连续亏损
            recent_years = [actual_end_year - 2, actual_end_year - 1, actual_end_year]
            consecutive_losses = 0
            
            for year in recent_years:
                if year >= start_year:  # 确保年份在我们的数据范围内
                    year_end = f"{year}1231"
                    try:
                        # 获取净利润数据进行预筛选
                        profit_check = self.token_manager.make_request(
                            lambda pro: pro.fina_indicator(
                                ts_code=stock_code,
                                end_date=year_end,
                                period_type='Y',
                                fields='ts_code,end_date,netprofit_margin'
                            )
                        )
                        if profit_check is not None and not profit_check.empty:
                            year_data = profit_check[profit_check['end_date'].str.startswith(str(year))]
                            if not year_data.empty:
                                net_margin = year_data.iloc[0]['netprofit_margin']
                                if net_margin is not None and net_margin < 0:
                                    consecutive_losses += 1
                        if self.use_delay:
                            time.sleep(0.05)  # 大幅减少延时：从0.1秒减少到0.05秒
                    except Exception as e:
                        logger.warning(f"预筛选检查失败 {stock_code} {year}: {e}")
                        # 如果预筛选失败，继续处理，不跳过
                        break
            
            # 如果最近3年连续亏损，跳过此股票
            if consecutive_losses >= 3:
                logger.info(f"跳过连续亏损股票 {stock_code}，连续亏损{consecutive_losses}年")
                return None
            
            data = {
                'financial_indicators': [],
                'balance_sheet': [],
                'dividend': [],
                'pe': [],
                'pb': [],
                'cashflow': []
            }
            
            # 获取每年的年报财务指标
            for year in years:
                # 获取年报财务指标（使用年末日期）
                year_end = f"{year}1231"
                
                # 1. 主要财务指标
                indicators = self.token_manager.make_request(
                    lambda pro: pro.fina_indicator(
                        ts_code=stock_code,
                        end_date=year_end,
                        period_type='Y',
                        fields='ts_code,end_date,roe,grossprofit_margin,netprofit_margin,debt_to_assets,current_ratio,assets_turn'
                    )
                )
                if indicators is not None and not indicators.empty:
                    # 过滤出该年的年报数据
                    year_indicators = indicators[indicators['end_date'].str.startswith(str(year))]
                    if not year_indicators.empty:
                        # 只取最新的一条年报数据
                        latest_indicator = year_indicators.iloc[0:1]
                        data['financial_indicators'].extend(latest_indicator.to_dict('records'))
                
                # 2. 资产负债表数据（获取营收等）
                balance_sheet = self.token_manager.make_request(
                    lambda pro: pro.balancesheet(
                        ts_code=stock_code,
                        end_date=year_end,
                        period_type='Y',
                        fields='ts_code,end_date,total_assets'
                    )
                )
                if balance_sheet is not None and not balance_sheet.empty:
                    year_balance = balance_sheet[balance_sheet['end_date'].str.startswith(str(year))]
                    if not year_balance.empty:
                        data['balance_sheet'].extend(year_balance.iloc[0:1].to_dict('records'))
                
                # 3. 现金流量表数据
                cashflow = self.token_manager.make_request(
                    lambda pro: pro.cashflow(
                        ts_code=stock_code,
                        end_date=year_end,
                        period_type='Y',
                        fields='ts_code,end_date,n_cashflow_act,net_profit'
                    )
                )
                if cashflow is not None and not cashflow.empty:
                    year_cashflow = cashflow[cashflow['end_date'].str.startswith(str(year))]
                    if not year_cashflow.empty:
                        data['cashflow'].extend(year_cashflow.iloc[0:1].to_dict('records'))
                
                # 4. 获取年末股息率（尝试多个日期）
                dividend_found = False
                for month_day in ['1231', '1230', '1229', '1228']:  # 尝试年末几天
                    test_date = f"{year}{month_day}"
                    dividend = self.token_manager.make_request(
                        lambda pro: pro.daily_basic(
                            ts_code=stock_code,
                            trade_date=test_date,
                            fields='ts_code,trade_date,dv_ratio'
                        )
                    )
                    if dividend is not None and not dividend.empty:
                        data['dividend'].extend(dividend.to_dict('records'))
                        dividend_found = True
                        break
                    if self.use_delay:
                        time.sleep(0.02)  # 大幅减少延时
                
                # 5. 获取年末PE（尝试多个日期）
                pe_found = False
                for month_day in ['1231', '1230', '1229', '1228']:  # 尝试年末几天
                    test_date = f"{year}{month_day}"
                    pe = self.token_manager.make_request(
                        lambda pro: pro.daily_basic(
                            ts_code=stock_code,
                            trade_date=test_date,
                            fields='ts_code,trade_date,pe'
                        )
                    )
                    if pe is not None and not pe.empty:
                        data['pe'].extend(pe.to_dict('records'))
                        pe_found = True
                        break
                    if self.use_delay:
                        time.sleep(0.02)  # 大幅减少延时
                
                # 6. 获取年末PB（尝试多个日期）
                pb_found = False
                for month_day in ['1231', '1230', '1229', '1228']:  # 尝试年末几天
                    test_date = f"{year}{month_day}"
                    pb = self.token_manager.make_request(
                        lambda pro: pro.daily_basic(
                            ts_code=stock_code,
                            trade_date=test_date,
                            fields='ts_code,trade_date,pb'
                        )
                    )
                    if pb is not None and not pb.empty:
                        data['pb'].extend(pb.to_dict('records'))
                        pb_found = True
                        break
                    if self.use_delay:
                        time.sleep(0.02)  # 大幅减少延时
                
                if self.use_delay:
                    time.sleep(0.1)  # 每年数据间隔：从0.3秒减少到0.1秒
            
            return data
            
        except Exception as e:
            logger.error(f"获取年度数据失败 {stock_code}: {e}")
            return None

    def process_batch(self, stocks_batch, start_year, end_year, use_cache=True):
        """处理一批股票数据"""
        batch_index = stocks_batch.index[0] // self.batch_size
        
        # 尝试从缓存加载
        if use_cache:
            cached_data = self._load_batch_from_cache(batch_index)
            if cached_data is not None:
                return cached_data
        
        # 获取新数据
        batch_data = {}
        for _, stock in stocks_batch.iterrows():
            stock_code = stock['ts_code']
            stock_data = self.get_annual_data(stock_code, start_year, end_year)
            if stock_data:
                batch_data[stock_code] = {
                    'name': stock['name'],
                    'industry': stock['industry'],
                    'data': stock_data
                }
        
        # 保存到缓存
        if batch_data:
            self._save_batch_to_cache(batch_data, batch_index)
        
        return batch_data

class ExcelOptimizer:
    """Excel数据优化器"""
    def __init__(self, df):
        self.df = df
        
    def create_summary_view(self):
        """创建汇总视图 - 只显示关键信息"""
        if self.df is None:
            return None
            
        # 基本信息列
        basic_cols = ['stock_code', 'stock_name', 'industry', 'need_analysis']
        
        # 计算各指标的最新值和平均值
        summary_data = []
        
        for _, row in self.df.iterrows():
            summary_row = {}
            
            # 基本信息
            for col in basic_cols:
                if col in row:
                    summary_row[col] = row[col]
            
            # ROE汇总
            roe_cols = [col for col in self.df.columns if col.startswith('roe_')]
            roe_values = [row[col] for col in roe_cols if pd.notna(row[col])]
            if roe_values:
                summary_row['roe_最新'] = roe_values[-1]
                summary_row['roe_平均'] = np.mean(roe_values)
                summary_row['roe_趋势'] = '上升' if len(roe_values) > 1 and roe_values[-1] > roe_values[0] else '下降'
            
            # 毛利率汇总
            gm_cols = [col for col in self.df.columns if col.startswith('gross_margin_')]
            gm_values = [row[col] for col in gm_cols if pd.notna(row[col])]
            if gm_values:
                summary_row['毛利率_最新'] = gm_values[-1]
                summary_row['毛利率_平均'] = np.mean(gm_values)
            
            # 净利率汇总
            nm_cols = [col for col in self.df.columns if col.startswith('net_margin_')]
            nm_values = [row[col] for col in nm_cols if pd.notna(row[col])]
            if nm_values:
                summary_row['净利率_最新'] = nm_values[-1]
                summary_row['净利率_平均'] = np.mean(nm_values)
            
            # PE汇总
            pe_cols = [col for col in self.df.columns if col.startswith('pe_')]
            pe_values = [row[col] for col in pe_cols if pd.notna(row[col])]
            if pe_values:
                summary_row['PE_最新'] = pe_values[-1]
                summary_row['PE_平均'] = np.mean(pe_values)
            
            # 股息率汇总
            div_cols = [col for col in self.df.columns if col.startswith('dividend_')]
            div_values = [row[col] for col in div_cols if pd.notna(row[col])]
            if div_values:
                summary_row['股息率_最新'] = div_values[-1]
                summary_row['股息率_平均'] = np.mean(div_values)
            
            # 综合评分（简单评分逻辑）
            score = 0
            if 'roe_平均' in summary_row and summary_row['roe_平均'] > 15:
                score += 20
            if '毛利率_平均' in summary_row and summary_row['毛利率_平均'] > 30:
                score += 20
            if '净利率_平均' in summary_row and summary_row['净利率_平均'] > 10:
                score += 20
            if 'PE_平均' in summary_row and 10 < summary_row['PE_平均'] < 25:
                score += 20
            if '股息率_平均' in summary_row and summary_row['股息率_平均'] > 2:
                score += 20
            
            summary_row['综合评分'] = score
            summary_data.append(summary_row)
        
        return pd.DataFrame(summary_data)
    
    def create_sector_analysis(self):
        """创建行业分析视图"""
        if self.df is None:
            return None
            
        # 按行业分组统计
        sector_stats = []
        
        for industry in self.df['industry'].unique():
            if pd.isna(industry):
                continue
                
            industry_data = self.df[self.df['industry'] == industry]
            
            # 计算行业平均指标
            roe_cols = [col for col in self.df.columns if col.startswith('roe_')]
            pe_cols = [col for col in self.df.columns if col.startswith('pe_')]
            
            sector_row = {
                '行业': industry,
                '公司数量': len(industry_data),
                '平均ROE': industry_data[roe_cols].mean().mean(),
                '平均PE': industry_data[pe_cols].mean().mean(),
                '高ROE公司数': (industry_data[roe_cols].mean(axis=1) > 15).sum(),
                '需要分析数': (industry_data['need_analysis'] == True).sum()
            }
            sector_stats.append(sector_row)
        
        return pd.DataFrame(sector_stats).sort_values('平均ROE', ascending=False)
    
    def create_filtered_views(self):
        """创建筛选视图"""
        if self.df is None:
            return {}
            
        views = {}
        
        # 高ROE股票（ROE均值>15%）
        roe_cols = [col for col in self.df.columns if col.startswith('roe_')]
        high_roe_mask = self.df[roe_cols].mean(axis=1) > 15
        views['高ROE股票'] = self.df[high_roe_mask][['stock_code', 'stock_name', 'industry'] + roe_cols]
        
        # 低PE股票（PE均值<20）
        pe_cols = [col for col in self.df.columns if col.startswith('pe_')]
        low_pe_mask = self.df[pe_cols].mean(axis=1) < 20
        views['低PE股票'] = self.df[low_pe_mask][['stock_code', 'stock_name', 'industry'] + pe_cols]
        
        # 高股息股票（股息率均值>3%）
        div_cols = [col for col in self.df.columns if col.startswith('dividend_')]
        high_div_mask = self.df[div_cols].mean(axis=1) > 3
        views['高股息股票'] = self.df[high_div_mask][['stock_code', 'stock_name', 'industry'] + div_cols]
        
        # 稳定盈利股票（净利率连续正值）
        nm_cols = [col for col in self.df.columns if col.startswith('net_margin_')]
        stable_profit_mask = (self.df[nm_cols] > 0).all(axis=1)
        views['稳定盈利股票'] = self.df[stable_profit_mask][['stock_code', 'stock_name', 'industry'] + nm_cols]
        
        return views
    
    def _apply_styles(self, excel_file):
        """应用Excel样式"""
        try:
            wb = load_workbook(excel_file)
            
            # 为每个工作表添加样式
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 设置标题行样式
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                
                # 设置数据行样式
                for row in range(2, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal="center")
                        
                        # 为数值类型的单元格设置条件格式
                        if isinstance(cell.value, (int, float)):
                            if cell.value is not None and cell.value < 0:
                                cell.font = Font(color="FF0000")  # 负值显示红色
                            elif cell.value is not None and cell.value > 20:
                                cell.font = Font(color="008000")  # 高值显示绿色
                
                # 自动调整列宽
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 20)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_file)
            
        except Exception as e:
            logger.error(f"应用样式失败: {e}")
    
    def generate_analysis_suggestions(self):
        """生成分析建议"""
        if self.df is None:
            return None
            
        suggestions = []
        
        # 1. 高价值股票推荐
        summary_df = self.create_summary_view()
        if summary_df is not None:
            high_score_stocks = summary_df[summary_df['综合评分'] >= 80].sort_values('综合评分', ascending=False)
            if not high_score_stocks.empty:
                suggestions.append("🌟 高价值股票推荐：")
                for _, stock in high_score_stocks.head(10).iterrows():
                    suggestions.append(f"  • {stock['stock_name']}({stock['stock_code']}) - 评分:{stock['综合评分']}")
        
        # 2. 行业机会
        sector_df = self.create_sector_analysis()
        if sector_df is not None:
            top_sectors = sector_df.head(5)
            suggestions.append("\n📈 优势行业：")
            for _, sector in top_sectors.iterrows():
                suggestions.append(f"  • {sector['行业']} - 平均ROE:{sector['平均ROE']:.2f}%")
        
        # 3. 筛选建议
        filtered_views = self.create_filtered_views()
        suggestions.append("\n🔍 筛选建议：")
        for view_name, view_df in filtered_views.items():
            suggestions.append(f"  • {view_name}: {len(view_df)}只股票")
        
        return '\n'.join(suggestions)
    
    def save_optimized_excel(self, output_file='stock_analysis_optimized.xlsx'):
        """保存优化后的Excel文件"""
        if self.df is None:
            logger.error("没有数据可保存")
            return False
            
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 1. 汇总视图
                summary_df = self.create_summary_view()
                if summary_df is not None:
                    summary_df.to_excel(writer, sheet_name='汇总视图', index=False)
                    logger.info(f"汇总视图已创建，包含{len(summary_df)}行数据")
                
                # 2. 行业分析
                sector_df = self.create_sector_analysis()
                if sector_df is not None:
                    sector_df.to_excel(writer, sheet_name='行业分析', index=False)
                    logger.info(f"行业分析已创建，包含{len(sector_df)}个行业")
                
                # 3. 筛选视图
                filtered_views = self.create_filtered_views()
                for view_name, view_df in filtered_views.items():
                    if not view_df.empty:
                        view_df.to_excel(writer, sheet_name=view_name, index=False)
                        logger.info(f"{view_name}已创建，包含{len(view_df)}只股票")
                
                # 4. 原始数据（可选）
                self.df.to_excel(writer, sheet_name='原始数据', index=False)
                logger.info("原始数据已保留")
            
            # 添加样式
            self._apply_styles(output_file)
            logger.info(f"优化后的Excel文件已保存: {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"保存Excel文件失败: {e}")
            return False

def setup_logger():
    """配置日志"""
    log_path = "logs"
    os.makedirs(log_path, exist_ok=True)
    logger.add(
        os.path.join(log_path, "data_collection_{time}.log"),
        rotation="500 MB",
        encoding="utf-8"
    )

def process_stock_data(batch_data):
    """处理股票数据为最终格式"""
    results = []
    for stock_code, stock_info in batch_data.items():
        row = {
            'stock_code': stock_code,
            'stock_name': stock_info['name'],
            'industry': stock_info['industry'],
            'need_analysis': False
        }
        
        # 处理财务指标
        for indicator in stock_info['data']['financial_indicators']:
            year = indicator['end_date'][:4]
            row[f'roe_{year}'] = indicator['roe']
            row[f'gross_margin_{year}'] = indicator['grossprofit_margin']
            row[f'net_margin_{year}'] = indicator['netprofit_margin']
            row[f'debt_ratio_{year}'] = indicator['debt_to_assets']
            row[f'current_ratio_{year}'] = indicator['current_ratio']
            row[f'asset_turnover_{year}'] = indicator['assets_turn']
        
        # 处理资产负债表数据
        for balance in stock_info['data']['balance_sheet']:
            year = balance['end_date'][:4]
            row[f'total_assets_{year}'] = balance['total_assets']
        
        # 处理现金流数据
        for cf in stock_info['data']['cashflow']:
            year = cf['end_date'][:4]
            # 计算现金流质量比率（经营现金流/净利润）
            if cf['n_cashflow_act'] and cf['net_profit'] and cf['net_profit'] != 0:
                row[f'ocf_to_profit_{year}'] = cf['n_cashflow_act'] / cf['net_profit']
        
        # 处理股息率
        for dividend in stock_info['data']['dividend']:
            year = dividend['trade_date'][:4]
            row[f'dividend_{year}'] = dividend['dv_ratio']
        
        # 处理PE
        for pe_data in stock_info['data']['pe']:
            year = pe_data['trade_date'][:4]
            row[f'pe_{year}'] = pe_data['pe']
        
        # 处理PB
        for pb_data in stock_info['data']['pb']:
            year = pb_data['trade_date'][:4]
            row[f'pb_{year}'] = pb_data['pb']
        
        results.append(row)
    
    return results

def create_sqlite_database(db_path='stock_analysis.db'):
    """创建SQLite数据库和表结构"""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # 创建股票基本信息表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS stocks (
            stock_code TEXT PRIMARY KEY,
            stock_name TEXT,
            industry TEXT,
            list_date TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 创建财务指标表（长格式，便于查询）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS financial_metrics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            stock_code TEXT,
            year INTEGER,
            metric_name TEXT,
            metric_value REAL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (stock_code) REFERENCES stocks (stock_code)
        )
    ''')
    
    # 创建索引提高查询性能
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_stock_year 
        ON financial_metrics (stock_code, year)
    ''')
    
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_metric_name 
        ON financial_metrics (metric_name)
    ''')
    
    conn.commit()
    conn.close()
    logger.info(f"SQLite数据库已创建: {db_path}")

def save_to_sqlite(data, db_path='stock_analysis.db'):
    """保存数据到SQLite数据库"""
    conn = sqlite3.connect(db_path)
    
    for _, row in data.iterrows():
        # 插入股票基本信息
        conn.execute('''
            INSERT OR REPLACE INTO stocks (stock_code, stock_name, industry)
            VALUES (?, ?, ?)
        ''', (row['stock_code'], row['stock_name'], row['industry']))
        
        # 插入财务指标数据
        for col in row.index:
            if col in ['stock_code', 'stock_name', 'industry', 'need_analysis']:
                continue
                
            # 解析指标名称和年份
            if '_' in col:
                parts = col.split('_')
                if len(parts) >= 2:
                    metric_name = '_'.join(parts[:-1])
                    year = parts[-1]
                    
                    if pd.notna(row[col]) and year.isdigit():
                        conn.execute('''
                            INSERT OR REPLACE INTO financial_metrics 
                            (stock_code, year, metric_name, metric_value)
                            VALUES (?, ?, ?, ?)
                        ''', (row['stock_code'], int(year), metric_name, float(row[col])))
    
    conn.commit()
    conn.close()
    logger.info(f"数据已保存到SQLite数据库: {db_path}")

def main():
    """主程序入口"""
    setup_logger()
    
    # 创建SQLite数据库
    create_sqlite_database()
    
    # 命令行参数解析
    parser = argparse.ArgumentParser(description='A股基本面数据收集工具 - 支持多Token')
    parser.add_argument('--limit', type=int, default=None, help='限制处理的股票数量（测试用）')
    parser.add_argument('--batch-size', type=int, default=50, help='批处理大小')
    parser.add_argument('--no-cache', action='store_true', help='不使用缓存，重新获取数据')
    parser.add_argument('--start-year', type=int, default=2019, help='开始年份')
    parser.add_argument('--end-year', type=int, default=2023, help='结束年份')
    parser.add_argument('--no-optimize', action='store_true', help='不生成优化Excel视图')
    parser.add_argument('--no-delay', action='store_true', help='不使用延时，最快速度运行（可能触发API限制）')
    parser.add_argument('--no-realtime-db', action='store_true', help='不实时更新数据库，仅在最后统一保存')
    
    args = parser.parse_args()
    
    # 从配置文件获取所有token
    tokens = []
    if config.TUSHARE_TOKENS:
        tokens = config.TUSHARE_TOKENS
    elif config.TUSHARE_TOKEN:
        tokens = [config.TUSHARE_TOKEN]
        
    if not tokens:
        logger.error("请在.env文件中设置TUSHARE_TOKEN或TUSHARE_TOKENS")
        return
    
    logger.info(f"🔧 配置的Token数量: {len(tokens)}")
    
    # 初始化收集器（传入所有tokens）
    collector = StockDataCollector(
        tokens,  # 传入所有tokens
        cache_dir='cache', 
        batch_size=args.batch_size,
        use_delay=not args.no_delay  # 如果指定了no_delay，则不使用延时
    )
    
    try:
        logger.info(f"数据收集时间范围：{args.start_year} 至 {args.end_year}")
        
        # 获取股票列表
        stocks = collector.get_all_stocks()
        if stocks is None:
            logger.error("获取股票列表失败")
            return
        
        # 限制股票数量（测试模式）
        if args.limit:
            stocks = stocks.head(args.limit)
            logger.info(f"限制模式：只处理前 {args.limit} 只股票")
        
        logger.info(f"共获取到 {len(stocks)} 只股票")
        
        # 计算批次数量
        total_batches = (len(stocks) + args.batch_size - 1) // args.batch_size
        logger.info(f"将分 {total_batches} 个批次处理")
        
        all_results = []
        
        # 按批次处理
        for i in tqdm(range(total_batches), desc="处理批次"):
            start_idx = i * args.batch_size
            end_idx = min((i + 1) * args.batch_size, len(stocks))
            stocks_batch = stocks.iloc[start_idx:end_idx]
            
            # 处理当前批次
            batch_data = collector.process_batch(
                stocks_batch,
                args.start_year,
                args.end_year,
                use_cache=not args.no_cache
            )
            
            # 处理数据并添加到结果
            if batch_data:
                batch_results = process_stock_data(batch_data)
                all_results.extend(batch_results)
                
                # 🔄 实时保存当前批次到数据库
                if batch_results:
                    batch_df = pd.DataFrame(batch_results)
                    if not args.no_realtime_db:
                        save_to_sqlite(batch_df)
                        logger.info(f"✅ 批次 {i+1} 数据已保存到数据库（{len(batch_results)}只股票）")
                    else:
                        logger.info(f"📦 批次 {i+1} 数据已缓存（{len(batch_results)}只股票），将在最后统一保存")
                
                logger.info(f"完成第 {i+1}/{total_batches} 批次处理，当前已处理 {len(all_results)} 只股票")
        
        # 保存最终结果
        if all_results:
            df = pd.DataFrame(all_results)
            output_file = 'stock_analysis_data.xlsx'
            df.to_excel(output_file, index=False)
            logger.info(f"原始数据已保存到: {output_file}")
            
            # 显示过滤效果统计
            total_attempted = len(stocks)
            successfully_processed = len(all_results)
            filtered_out = total_attempted - successfully_processed
            filter_rate = (filtered_out / total_attempted * 100) if total_attempted > 0 else 0
            
            # 显示Token使用统计
            token_stats = collector.token_manager.get_stats()
            
            logger.info(f"📊 数据处理统计:")
            logger.info(f"  • 总股票数: {total_attempted}")
            logger.info(f"  • 成功处理: {successfully_processed}")
            logger.info(f"  • 过滤掉数: {filtered_out} ({filter_rate:.1f}%)")
            logger.info(f"  • 数据列数: {len(df.columns)}")
            
            logger.info(f"🔧 Token使用统计:")
            logger.info(f"  • 总请求数: {token_stats['total_requests']}")
            logger.info(f"  • 成功请求: {token_stats['successful_requests']}")
            logger.info(f"  • 成功率: {token_stats['success_rate']}")
            logger.info(f"  • 当前Token: {token_stats['current_token']}/{token_stats['total_tokens']}")
            
            # 自动生成优化视图
            if not args.no_optimize:
                logger.info("开始生成优化Excel视图...")
                optimizer = ExcelOptimizer(df)
                
                if optimizer.save_optimized_excel():
                    logger.info("✅ 优化Excel文件创建成功: stock_analysis_optimized.xlsx")
                    
                    # 生成分析建议
                    suggestions = optimizer.generate_analysis_suggestions()
                    if suggestions:
                        logger.info("📋 投资分析建议：")
                        print("\n" + suggestions)
                        
                        # 保存建议到文件
                        with open('analysis_suggestions.txt', 'w', encoding='utf-8') as f:
                            f.write(suggestions)
                        logger.info("💾 分析建议已保存到: analysis_suggestions.txt")
                    
                    print(f"\n🎯 数据处理完成！生成了以下文件：")
                    print(f"  📄 {output_file} - 原始数据 ({len(df.columns)}列)")
                    print(f"  📊 stock_analysis_optimized.xlsx - 优化视图 (7个工作表)")
                    print(f"  📝 analysis_suggestions.txt - 投资建议")
                    print(f"\n🔧 Token统计: {token_stats['success_rate']} 成功率，使用了 {token_stats['total_tokens']} 个Token")
                else:
                    logger.error("优化Excel文件创建失败")
            else:
                logger.info("已跳过优化Excel视图生成（使用--no-optimize参数）")
                
            # 数据库保存逻辑
            if args.no_realtime_db:
                # 统一保存所有数据到数据库
                save_to_sqlite(df)
                logger.info("📊 所有数据已统一保存到SQLite数据库")
            else:
                # 数据已在批次处理时实时保存到数据库
                logger.info("📊 所有批次数据已实时保存到SQLite数据库")
            
        else:
            logger.error("没有收集到任何数据")
            
    except Exception as e:
        logger.error(f"主程序执行失败: {e}")
        # 显示Token统计（即使出错也显示）
        try:
            token_stats = collector.token_manager.get_stats()
            logger.info(f"🔧 最终Token统计: {token_stats}")
        except:
            pass

if __name__ == "__main__":
    main() 