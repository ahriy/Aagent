#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
价值投资Agent - 融合巴菲特、查理芒格、格雷厄姆的投资智慧
支持实时PE获取，减少API调用的智能筛选，并集成DeepSeek AI分析
"""

import sqlite3
import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from loguru import logger
import argparse
import json
import time
import random
import requests

# Tushare API相关导入
try:
    import tushare as ts
    TUSHARE_AVAILABLE = True
except ImportError:
    TUSHARE_AVAILABLE = False
    logger.warning("Tushare未安装，将无法获取实时PE数据")

class DeepSeekAnalyzer:
    """DeepSeek AI分析器"""
    
    def __init__(self, api_key: str = None, base_url: str = "https://api.deepseek.com", model: str = "deepseek-chat"):
        self.api_key = api_key
        self.base_url = base_url
        self.model = model
        self.headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.api_key}"
        } if api_key else None
        
        # 加载系统提示词
        self.system_prompt = self.load_system_prompt()
    
    def load_system_prompt(self) -> str:
        """加载系统提示词"""
        try:
            with open('system_prompt.md', 'r', encoding='utf-8') as f:
                return f.read()
        except FileNotFoundError:
            logger.warning("未找到system_prompt.md文件，使用默认提示词")
            return "你是一位专业的价值投资分析师，请对提供的股票数据进行深度分析。"
    
    def analyze_stock(self, stock_code: str, stock_data: Dict, score_details: Dict) -> Optional[str]:
        """使用DeepSeek分析单个股票"""
        if not self.api_key or not self.headers:
            logger.warning("DeepSeek API Key未配置，跳过AI分析")
            return None
        
        try:
            # 构建分析用的数据摘要
            data_summary = self.format_stock_data(stock_code, stock_data, score_details)
            
            # 构建请求
            payload = {
                "model": self.model,
                "messages": [
                    {
                        "role": "system",
                        "content": self.system_prompt
                    },
                    {
                        "role": "user",
                        "content": f"请对以下A股上市公司进行深度价值投资分析：\n\n{data_summary}"
                    }
                ],
                "temperature": 0.3,
                "max_tokens": 3000
            }
            
            # 发送请求
            api_url = self.base_url
            if not api_url.endswith('/chat/completions'):
                api_url = api_url.rstrip('/') + '/chat/completions'
            
            response = requests.post(
                api_url,
                headers=self.headers,
                json=payload,
                timeout=60
            )
            
            if response.status_code == 200:
                result = response.json()
                analysis = result['choices'][0]['message']['content']
                logger.info(f"完成{stock_code}的DeepSeek AI分析")
                return analysis
            else:
                logger.error(f"DeepSeek API请求失败: {response.status_code} - {response.text}")
                return None
                
        except Exception as e:
            logger.error(f"DeepSeek分析{stock_code}时出错: {e}")
            return None
    
    def format_stock_data(self, stock_code: str, stock_data: Dict, score_details: Dict) -> str:
        """格式化股票数据为分析用的文本"""
        basic_info = stock_data.get('basic_info', {})
        metrics = stock_data.get('metrics', {})
        
        # 基本信息
        company_name = basic_info.get('stock_name', '未知')
        industry = basic_info.get('industry', '未知')
        
        # 评分信息
        total_score = score_details.get('total_score', 0)
        buffett_score = score_details.get('buffett', {}).get('score', 0)
        munger_score = score_details.get('munger', {}).get('score', 0)
        graham_score = score_details.get('graham', {}).get('score', 0)
        
        # 关键财务指标
        data_text = f"""
股票代码：{stock_code}
公司名称：{company_name}
所属行业：{industry}

【价值投资评分】
总分：{total_score:.1f}/100
- 巴菲特策略得分：{buffett_score:.1f}/100
- 芒格策略得分：{munger_score:.1f}/100  
- 格雷厄姆策略得分：{graham_score:.1f}/100

【关键财务指标】（近5年数据）
"""
        
        # 添加重要的财务指标
        key_metrics = [
            ('净资产收益率(%)', 'roe'),
            ('市盈率', 'pe_ratio'),
            ('市净率', 'pb_ratio'),
            ('净利润(万元)', 'net_profit'),
            ('营业收入(万元)', 'operating_income'),
            ('资产负债率(%)', 'debt_to_asset_ratio'),
            ('流动比率', 'current_ratio'),
            ('速动比率', 'quick_ratio'),
            ('每股收益(元)', 'eps'),
            ('每股净资产(元)', 'book_value_per_share'),
            ('毛利率(%)', 'gross_profit_margin'),
            ('净利率(%)', 'net_profit_margin')
        ]
        
        for metric_name, metric_key in key_metrics:
            if metric_key in metrics:
                metric_data = metrics[metric_key]
                values = []
                for year in sorted(metric_data.keys(), reverse=True)[:5]:  # 最近5年
                    value = metric_data[year]
                    if pd.notna(value):
                        values.append(f"{year}年: {value}")
                
                if values:
                    data_text += f"\n{metric_name}：{' | '.join(values)}"
        
        # 添加评分详情
        data_text += f"\n\n【详细评分分析】"
        
        for strategy, details in score_details.items():
            if strategy in ['buffett', 'munger', 'graham'] and isinstance(details, dict):
                data_text += f"\n\n{strategy.title()}策略评分详情："
                for criterion, score in details.get('scores', {}).items():
                    data_text += f"\n- {criterion}: {score}分"
        
        return data_text

class TushareManager:
    """Tushare API管理器 - 支持多token和智能重试"""
    
    def __init__(self, tokens: List[str]):
        if not TUSHARE_AVAILABLE:
            raise ImportError("需要安装tushare: pip install tushare")
        
        self.tokens = tokens
        self.current_token_index = 0
        self.pro = None
        self.init_api()
        
    def init_api(self):
        """初始化API连接"""
        if self.tokens:
            current_token = self.tokens[self.current_token_index]
            self.pro = ts.pro_api(current_token)
            logger.info(f"使用Tushare Token {self.current_token_index + 1}/{len(self.tokens)}")
    
    def switch_token(self):
        """切换到下一个token"""
        if len(self.tokens) > 1:
            self.current_token_index = (self.current_token_index + 1) % len(self.tokens)
            self.init_api()
            logger.info(f"切换到Token {self.current_token_index + 1}")
            return True
        return False
    
    def get_realtime_pe(self, ts_code: str, max_retries: int = 3) -> Optional[float]:
        """获取实时PE数据"""
        if not self.pro:
            return None
            
        for attempt in range(max_retries):
            try:
                # 获取基本信息，包含PE_TTM
                df = self.pro.daily_basic(ts_code=ts_code, trade_date='', fields='ts_code,pe_ttm')
                
                if not df.empty and pd.notna(df.iloc[0]['pe_ttm']):
                    pe = float(df.iloc[0]['pe_ttm'])
                    if pe > 0:  # 确保PE为正数
                        logger.debug(f"获取{ts_code}实时PE: {pe:.2f}")
                        return pe
                
                return None
                
            except Exception as e:
                logger.warning(f"获取{ts_code}实时PE失败 (尝试{attempt+1}/{max_retries}): {e}")
                
                # 如果是API限制相关错误，尝试切换token
                if any(keyword in str(e).lower() for keyword in ['limit', 'timeout', 'rate', '限制']):
                    if self.switch_token():
                        time.sleep(1)  # 切换后稍作等待
                        continue
                
                # 否则等待后重试
                if attempt < max_retries - 1:
                    wait_time = (attempt + 1) * 2 + random.uniform(0.5, 1.5)
                    time.sleep(wait_time)
        
        return None

class ValueInvestmentAgent:
    """
    价值投资Agent
    融合三位投资大师的理念：
    - 巴菲特：寻找护城河、长期价值、合理价格
    - 查理芒格：理性思维、质量优先、反向思考
    - 格雷厄姆：安全边际、价值与价格差异、基本面分析
    
    支持实时PE获取、智能筛选和DeepSeek AI深度分析
    """
    
    def __init__(self, db_path='stock_analysis.db', tushare_config_path='config.json'):
        self.db_path = db_path
        self.tushare_manager = None
        self.deepseek_analyzer = None
        self.setup_logger()
        self.load_tushare_config(tushare_config_path)
        self.load_deepseek_config(tushare_config_path)
        
    def setup_logger(self):
        """设置日志"""
        logger.add(
            "logs/value_investment_{time}.log",
            rotation="10 MB",
            encoding="utf-8"
        )
    
    def load_tushare_config(self, config_path: str):
        """加载Tushare配置"""
        tokens = []
        
        # 尝试从JSON配置文件加载
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                tokens = config.get('tushare_tokens', [])
        except FileNotFoundError:
            logger.debug(f"未找到JSON配置文件{config_path}")
        except Exception as e:
            logger.warning(f"读取JSON配置文件失败: {e}")
        
        # 如果JSON配置为空，尝试从config.py加载
        if not tokens:
            try:
                import config
                tokens = getattr(config, 'TUSHARE_TOKENS', [])
                if tokens:
                    logger.info("从config.py加载Tushare配置")
            except ImportError:
                logger.debug("未找到config.py模块")
            except Exception as e:
                logger.warning(f"从config.py加载配置失败: {e}")
        
        if tokens and TUSHARE_AVAILABLE:
            self.tushare_manager = TushareManager(tokens)
            logger.info(f"已加载{len(tokens)}个Tushare token")
        else:
            logger.warning("未配置Tushare tokens或tushare不可用，将使用历史PE数据")
    
    def load_deepseek_config(self, config_path: str):
        """加载DeepSeek配置"""
        deepseek_key = None
        api_url = "https://api.deepseek.com"  # 默认URL
        model_name = "deepseek-ai/DeepSeek-R1"  # 默认模型名称
        
        # 尝试从JSON配置文件加载
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                deepseek_key = config.get('deepseek_api_key')
                api_url = config.get('api_url', api_url)
                model_name = config.get('model_name', model_name)
        except FileNotFoundError:
            logger.debug(f"未找到JSON配置文件{config_path}")
        except Exception as e:
            logger.warning(f"读取JSON配置文件失败: {e}")
        
        # 如果JSON配置为空，尝试从config.py加载
        if not deepseek_key:
            try:
                import config
                deepseek_key = getattr(config, 'DEEPSEEK_API_KEY', '')
                api_url = getattr(config, 'API_URL', api_url)
                model_name = getattr(config, 'DEEPSEEK_MODEL', model_name)
                if deepseek_key:
                    logger.info("从config.py加载DeepSeek配置")
            except ImportError:
                logger.debug("未找到config.py模块")
            except Exception as e:
                logger.warning(f"从config.py加载配置失败: {e}")
        
        if deepseek_key:
            self.deepseek_analyzer = DeepSeekAnalyzer(api_key=deepseek_key, base_url=api_url, model=model_name)
            logger.info(f"已加载DeepSeek API配置 (URL: {api_url}, Model: {model_name})")
        else:
            logger.warning("未配置DeepSeek API key，将跳过AI分析")
    
    def normalize_stock_code(self, stock_code: str) -> str:
        """标准化股票代码格式，支持数据库查询"""
        # 如果已经是完整格式，直接返回
        if '.' in stock_code:
            return stock_code
        
        # 6位代码，添加交易所后缀
        if len(stock_code) == 6:
            if stock_code.startswith(('60', '68')):
                return f"{stock_code}.SH"
            elif stock_code.startswith(('00', '30')):
                return f"{stock_code}.SZ"
        
        return stock_code
    
    def get_realtime_pe(self, stock_code: str) -> Optional[float]:
        """获取实时PE数据"""
        if not self.tushare_manager:
            return None
            
        ts_code = self.normalize_stock_code(stock_code)
        return self.tushare_manager.get_realtime_pe(ts_code)
    
    def _should_skip_stock(self, stock_code: str, stock_name: str) -> Tuple[bool, str]:
        """
        判断是否应该跳过此股票
        
        Args:
            stock_code: 股票代码
            stock_name: 股票名称
        
        Returns:
            (是否跳过, 跳过原因)
        """
        # 1. 跳过ST股票
        if stock_name and ('ST' in stock_name or '*ST' in stock_name or 'PT' in stock_name):
            return True, "ST股票风险过高"
        
        # 2. 跳过B股
        if stock_code.endswith('.BJ'):  # 北交所
            return True, "北交所股票流动性较低"
        
        # 3. 检查连续亏损
        conn = sqlite3.connect(self.db_path)
        try:
            # 获取最近3年的净利润数据
            query = """
            SELECT year, metric_value 
            FROM financial_metrics 
            WHERE stock_code = ? AND metric_name = 'net_profit'
            ORDER BY year DESC
            LIMIT 3
            """
            df = pd.read_sql_query(query, conn, params=[stock_code])
            
            if len(df) >= 3:
                # 检查是否连续3年亏损
                recent_profits = df['metric_value'].tolist()
                if all(profit <= 0 for profit in recent_profits if pd.notna(profit)):
                    return True, "连续3年亏损"
            
            # 4. 检查是否有基本财务数据
            basic_metrics_query = """
            SELECT COUNT(DISTINCT metric_name) as metric_count
            FROM financial_metrics 
            WHERE stock_code = ? AND metric_name IN ('roe', 'pe', 'pb')
            """
            metric_count_df = pd.read_sql_query(basic_metrics_query, conn, params=[stock_code])
            
            if metric_count_df.iloc[0]['metric_count'] < 2:  # 至少要有2个基本指标
                return True, "缺乏基本财务数据"
            
            # 5. 检查ROE数据质量
            roe_query = """
            SELECT metric_value 
            FROM financial_metrics 
            WHERE stock_code = ? AND metric_name = 'roe'
            ORDER BY year DESC
            LIMIT 3
            """
            roe_df = pd.read_sql_query(roe_query, conn, params=[stock_code])
            
            if len(roe_df) >= 2:
                # 如果最近2年ROE都是负数，跳过
                recent_roe = roe_df['metric_value'].tolist()
                valid_roe = [roe for roe in recent_roe if pd.notna(roe)]
                if len(valid_roe) >= 2 and all(roe < 0 for roe in valid_roe):
                    return True, "近年ROE持续为负"
            
            return False, ""
            
        except Exception as e:
            logger.warning(f"筛选股票{stock_code}时出错: {e}")
            return False, ""
        finally:
            conn.close()
    
    def get_stock_metrics(self, stock_code: str, years: List[int] = None) -> Dict:
        """获取单个股票的财务指标"""
        if years is None:
            years = [2020, 2021, 2022, 2023, 2024]
        
        # 标准化股票代码
        normalized_code = self.normalize_stock_code(stock_code)
        
        conn = sqlite3.connect(self.db_path)
        
        # 获取基本信息
        basic_info = pd.read_sql_query(
            "SELECT * FROM stocks WHERE stock_code = ?", 
            conn, params=[normalized_code]
        )
        
        if basic_info.empty:
            conn.close()
            return {}
        
        # 获取财务指标
        metrics_data = pd.read_sql_query(
            "SELECT year, metric_name, metric_value FROM financial_metrics WHERE stock_code = ?",
            conn, params=[normalized_code]
        )
        
        conn.close()
        
        if metrics_data.empty:
            return {}
        
        # 重构数据格式
        result = {
            'basic_info': basic_info.iloc[0].to_dict(),
            'metrics': {}
        }
        
        for _, row in metrics_data.iterrows():
            year = int(row['year'])
            metric = row['metric_name']
            value = row['metric_value']
            
            if metric not in result['metrics']:
                result['metrics'][metric] = {}
            result['metrics'][metric][year] = value
            
        return result
    
    def buffett_criteria(self, stock_data: Dict, include_pe_evaluation: bool = False, realtime_pe: Optional[float] = None) -> Dict:
        """
        巴菲特选股标准
        - 持续高ROE (>15%)
        - 稳定盈利增长
        - 低债务比率 (<0.3)
        - 合理估值
        
        Args:
            stock_data: 股票数据
            include_pe_evaluation: 是否包含PE评估
            realtime_pe: 实时PE数据（优先使用）
        """
        score = 0
        details = []
        
        metrics = stock_data.get('metrics', {})
        
        # 1. ROE分析 - 巴菲特最看重的指标
        roe_data = metrics.get('roe', {})
        if roe_data:
            roe_values = [v for v in roe_data.values() if v is not None and v > 0]
            if len(roe_values) >= 3:
                avg_roe = np.mean(roe_values)  # 保持百分比形式用于显示
                # 转换为小数进行比较
                avg_roe_decimal = avg_roe / 100.0
                if avg_roe_decimal > 0.20:
                    score += 25
                    details.append(f"🌟 卓越ROE: {avg_roe:.1f}% (>20%)")
                elif avg_roe_decimal > 0.15:
                    score += 15
                    details.append(f"✅ 优秀ROE: {avg_roe:.1f}% (>15%)")
                elif avg_roe_decimal > 0.10:
                    score += 5
                    details.append(f"📊 一般ROE: {avg_roe:.1f}% (>10%)")
                
                # ROE稳定性
                roe_std = np.std(roe_values)
                if roe_std < 5:
                    score += 10
                    details.append(f"🎯 ROE稳定性良好 (标准差: {roe_std:.1f})")
        
        # 2. 债务比率分析
        debt_data = metrics.get('debt_ratio', {})
        if debt_data:
            debt_values = [v for v in debt_data.values() if v is not None]
            if debt_values:
                avg_debt = np.mean(debt_values) / 100.0
                if avg_debt < 0.2:
                    score += 20
                    details.append(f"💪 低债务负担: {avg_debt:.1%} (<20%)")
                elif avg_debt < 0.3:
                    score += 10
                    details.append(f"✅ 适度债务: {avg_debt:.1%} (<30%)")
                elif avg_debt > 0.6:
                    score -= 10
                    details.append(f"⚠️ 高债务风险: {avg_debt:.1%} (>60%)")
        
        # 3. 盈利稳定性（净利率）
        net_margin_data = metrics.get('net_margin', {})
        if net_margin_data:
            margin_values = [v for v in net_margin_data.values() if v is not None]
            positive_margins = [v for v in margin_values if v > 0]
            
            if len(positive_margins) == len(margin_values) and len(margin_values) >= 3:
                score += 15
                avg_margin = np.mean(positive_margins)
                details.append(f"📈 持续盈利: 净利率 {avg_margin:.1f}%")
        
        # 4. 流动性分析
        current_ratio_data = metrics.get('current_ratio', {})
        if current_ratio_data:
            cr_values = [v for v in current_ratio_data.values() if v is not None]
            if cr_values:
                avg_cr = np.mean(cr_values)
                if avg_cr > 2:
                    score += 10
                    details.append(f"💰 流动性充裕: {avg_cr:.1f}")
                elif avg_cr > 1.5:
                    score += 5
                    details.append(f"✅ 流动性良好: {avg_cr:.1f}")
        
        # 5. PE估值分析（可选）
        if include_pe_evaluation:
            current_pe = realtime_pe
            pe_source = "实时"
            
            # 如果没有实时PE，使用历史PE
            if current_pe is None:
                pe_data = metrics.get('pe', {})
                if pe_data:
                    latest_year = max(pe_data.keys())
                    current_pe = pe_data[latest_year]
                    pe_source = "历史"
            
            if current_pe is not None and 0 < current_pe < 100:
                if current_pe < 15:
                    score += 25
                    details.append(f"💎 低估值: PE {current_pe:.1f}x (<15, {pe_source})")
                elif current_pe < 25:
                    score += 15
                    details.append(f"✅ 合理估值: PE {current_pe:.1f}x (<25, {pe_source})")
                elif current_pe < 35:
                    score += 5
                    details.append(f"📊 适中估值: PE {current_pe:.1f}x (<35, {pe_source})")
                elif current_pe > 50:
                    score -= 10
                    details.append(f"⚠️ 估值偏高: PE {current_pe:.1f}x (>50, {pe_source})")
        
        return {
            'score': min(score, 100),
            'details': details,
            'methodology': '巴菲特标准：护城河、持续盈利、低债务' + ('、合理估值' if include_pe_evaluation else '')
        }
    
    def munger_criteria(self, stock_data: Dict, include_pe_evaluation: bool = False, realtime_pe: Optional[float] = None) -> Dict:
        """
        芒格选股标准
        - 简单易懂的生意模式
        - 质量优于价格
        - 理性分析，避免情绪化
        
        Args:
            stock_data: 股票数据
            include_pe_evaluation: 是否包含PE评估
            realtime_pe: 实时PE数据（优先使用）
        """
        score = 0
        details = []
        
        metrics = stock_data.get('metrics', {})
        industry = stock_data.get('basic_info', {}).get('industry', '')
        
        # 1. 行业质量评估（芒格偏好的行业）
        quality_industries = [
            '银行', '保险', '食品饮料', '白酒', '医药生物', '公用事业',
            '消费服务', '商业贸易', '家用电器', '餐饮', '乳制品', 
            '调味品', '软饮料', '酿酒', '中药', '生物制药', '医疗器械',
            '电力', '燃气', '水务', '机场', '高速公路', '港口'
        ]
        
        if any(keyword in industry for keyword in quality_industries):
            score += 15
            details.append(f"🎯 优质行业: {industry}")
        
        # 2. 资产周转率（经营效率）
        asset_turnover_data = metrics.get('asset_turnover', {})
        if asset_turnover_data:
            at_values = [v for v in asset_turnover_data.values() if v is not None and v > 0]
            if at_values:
                avg_turnover = np.mean(at_values)
                if avg_turnover > 0.8:
                    score += 15
                    details.append(f"⚡ 高效运营: 资产周转率 {avg_turnover:.2f}")
                elif avg_turnover > 0.5:
                    score += 10
                    details.append(f"✅ 运营良好: 资产周转率 {avg_turnover:.2f}")
        
        # 3. 毛利率稳定性（产品定价权）
        gross_margin_data = metrics.get('gross_margin', {})
        if gross_margin_data:
            gm_values = [v for v in gross_margin_data.values() if v is not None and v > 0]
            if len(gm_values) >= 3:
                avg_gm = np.mean(gm_values)  # 保持百分比形式用于显示
                gm_trend = self._calculate_trend(gm_values)
                
                # 转换为小数进行比较
                avg_gm_decimal = avg_gm / 100.0
                if avg_gm_decimal > 0.40:
                    score += 20
                    details.append(f"💎 高毛利率: {avg_gm:.1f}% (强定价权)")
                elif avg_gm_decimal > 0.25:
                    score += 10
                    details.append(f"✅ 良好毛利率: {avg_gm:.1f}%")
                
                if gm_trend > 0:
                    score += 10
                    details.append("📈 毛利率呈上升趋势")
        
        # 4. 反向思维：避开高估值陷阱（可选）
        if include_pe_evaluation:
            current_pe = realtime_pe
            pe_source = "实时"
            
            # 如果没有实时PE，使用历史PE
            if current_pe is None:
                pe_data = metrics.get('pe', {})
                if pe_data:
                    latest_year = max(pe_data.keys())
                    current_pe = pe_data[latest_year]
                    pe_source = "历史"
            
            if current_pe is not None and 0 < current_pe < 100:
                if current_pe < 20:
                    score += 15
                    details.append(f"💰 合理估值: PE {current_pe:.1f}x (<20, {pe_source})")
                elif current_pe < 30:
                    score += 5
                    details.append(f"📊 适中估值: PE {current_pe:.1f}x (<30, {pe_source})")
                elif current_pe > 60:
                    score -= 10
                    details.append(f"⚠️ 估值偏高: PE {current_pe:.1f}x (>60, {pe_source})")
        
        return {
            'score': min(score, 100),
            'details': details,
            'methodology': '芒格标准：质量优先、理性分析、长期视角' + ('、反向思维' if include_pe_evaluation else '')
        }
    
    def graham_criteria(self, stock_data: Dict, include_pe_evaluation: bool = True, realtime_pe: Optional[float] = None) -> Dict:
        """
        格雷厄姆选股标准
        - 安全边际
        - 价值投资的鼻祖理念
        - 重视资产负债表
        
        Args:
            stock_data: 股票数据
            include_pe_evaluation: 是否包含PE评估（格雷厄姆默认包含）
            realtime_pe: 实时PE数据（优先使用）
        """
        score = 0
        details = []
        
        metrics = stock_data.get('metrics', {})
        
        # 1. PE比率（格雷厄姆经典指标）
        if include_pe_evaluation:
            current_pe = realtime_pe
            pe_source = "实时"
            
            # 如果没有实时PE，使用历史PE
            if current_pe is None:
                pe_data = metrics.get('pe', {})
                if pe_data:
                    latest_year = max(pe_data.keys())
                    current_pe = pe_data[latest_year]
                    pe_source = "历史"
            
            if current_pe is not None and 0 < current_pe < 100:
                if current_pe < 12:
                    score += 25
                    details.append(f"🎯 低估值: PE {current_pe:.1f}x (<12, {pe_source})")
                elif current_pe < 20:
                    score += 15
                    details.append(f"✅ 合理估值: PE {current_pe:.1f}x (<20, {pe_source})")
                elif current_pe < 30:
                    score += 5
                    details.append(f"📊 适中估值: PE {current_pe:.1f}x (<30, {pe_source})")
                elif current_pe > 40:
                    score -= 5
                    details.append(f"⚠️ 估值偏高: PE {current_pe:.1f}x (>40, {pe_source})")
        
        # 2. PB比率（资产价值）
        pb_data = metrics.get('pb', {})
        if pb_data:
            # 使用最新年份的PB数据
            latest_year = max(pb_data.keys())
            current_pb = pb_data[latest_year]
            
            if current_pb is not None and current_pb > 0:
                if current_pb < 1:
                    score += 20
                    details.append(f"💎 破净股: PB {current_pb:.2f}x (<1)")
                elif current_pb < 2:
                    score += 10
                    details.append(f"✅ 低PB: {current_pb:.2f}x (<2)")
                elif current_pb < 3:
                    score += 5
                    details.append(f"📊 合理PB: {current_pb:.2f}x (<3)")
        
        # 3. 股息率（价值回报）
        dividend_data = metrics.get('dividend', {})
        if dividend_data:
            # 使用最新年份的股息率数据
            latest_year = max(dividend_data.keys())
            current_dividend = dividend_data[latest_year]
            
            if current_dividend is not None and current_dividend > 0:
                # 转换为小数进行比较
                current_dividend_decimal = current_dividend / 100.0
                if current_dividend_decimal > 0.04:
                    score += 15
                    details.append(f"💰 高股息: {current_dividend:.1f}% (>4%)")
                elif current_dividend_decimal > 0.02:
                    score += 10
                    details.append(f"✅ 良好股息: {current_dividend:.1f}% (>2%)")
        
        # 4. 流动比率（安全边际）
        current_ratio_data = metrics.get('current_ratio', {})
        if current_ratio_data:
            cr_values = [v for v in current_ratio_data.values() if v is not None]
            if cr_values:
                avg_cr = np.mean(cr_values)
                if avg_cr > 2:
                    score += 15
                    details.append(f"🛡️ 安全边际高: 流动比率 {avg_cr:.1f}")
                elif avg_cr > 1.5:
                    score += 10
                    details.append(f"✅ 安全边际适中: 流动比率 {avg_cr:.1f}")
        
        # 5. 总资产增长（企业发展）
        total_assets_data = metrics.get('total_assets', {})
        if total_assets_data:
            asset_values = [v for v in total_assets_data.values() if v is not None and v > 0]
            if len(asset_values) >= 3:
                asset_growth = self._calculate_growth_rate(asset_values)
                if asset_growth > 0.1:
                    score += 10
                    details.append(f"📈 资产稳健增长: {asset_growth:.1%}")
                elif asset_growth > 0:
                    score += 5
                    details.append(f"✅ 资产正增长: {asset_growth:.1%}")
        
        return {
            'score': min(score, 100),
            'details': details,
            'methodology': '格雷厄姆标准：安全边际、价值发现、基本面分析' + ('、PE估值' if include_pe_evaluation else '')
        }
    
    def preliminary_screening(self, stock_code: str) -> Dict:
        """
        基于历史数据的初步筛选评分（不使用实时PE）
        主要用于第一阶段快速筛选
        """
        try:
            stock_data = self.get_stock_metrics(stock_code)
            
            if not stock_data:
                return {'error': f'无法获取股票 {stock_code} 的数据'}
            
            metrics = stock_data.get('metrics', {})
            preliminary_score = 0
            
            # ROE评分 (15分)
            roe_data = metrics.get('roe', {})
            if roe_data:
                roe_values = [v for v in roe_data.values() if v is not None]
                if roe_values:
                    roe_avg = np.mean(roe_values) / 100.0  # 转换为小数
                    if roe_avg >= 0.15:  # ≥15%
                        preliminary_score += 15
                    elif roe_avg >= 0.10:  # 10-15%
                        preliminary_score += 10
                    elif roe_avg >= 0.05:  # 5-10%
                        preliminary_score += 5
            
            # 现金流评分 (10分)
            cf_data = metrics.get('operating_cash_flow', {})
            if cf_data:
                cf_count = sum(1 for v in cf_data.values() if v and v > 0)
                if cf_count >= 4:  # 4年都为正
                    preliminary_score += 10
                elif cf_count >= 3:  # 3年为正
                    preliminary_score += 7
                elif cf_count >= 2:  # 2年为正
                    preliminary_score += 3
            
            # 营收增长评分 (10分)
            revenue_data = metrics.get('revenue', {})
            avg_growth = None
            if revenue_data:
                revenue_values = [v for v in revenue_data.values() if v is not None and v > 0]
                if len(revenue_values) >= 3:
                    growth_rates = []
                    for i in range(1, len(revenue_values)):
                        if revenue_values[i-1] > 0:
                            growth = (revenue_values[i] - revenue_values[i-1]) / revenue_values[i-1]
                            growth_rates.append(growth)
                    
                    if growth_rates:
                        avg_growth = np.mean(growth_rates)
                        if avg_growth >= 0.10:  # ≥10%
                            preliminary_score += 10
                        elif avg_growth >= 0.05:  # 5-10%
                            preliminary_score += 7
                        elif avg_growth >= 0:  # 正增长
                            preliminary_score += 3
            
            # 财务稳定性评分 (10分)
            debt_data = metrics.get('debt_ratio', {})
            current_data = metrics.get('current_ratio', {})
            
            debt_ratio = None
            current_ratio = None
            
            if debt_data:
                debt_values = [v for v in debt_data.values() if v is not None]
                if debt_values:
                    debt_ratio = np.mean(debt_values) / 100.0  # 转换为小数
                    if debt_ratio < 0.3:  # 负债率<30%
                        preliminary_score += 5
                    elif debt_ratio < 0.5:  # 负债率<50%
                        preliminary_score += 3
            
            if current_data:
                current_values = [v for v in current_data.values() if v is not None]
                if current_values:
                    current_ratio = np.mean(current_values)
                    if current_ratio > 1.5:  # 流动比率>1.5
                        preliminary_score += 5
                    elif current_ratio > 1.0:  # 流动比率>1.0
                        preliminary_score += 3
            
            # 历史PE评分 (使用数据库中的历史PE，不调用API)
            pe_data = metrics.get('pe', {})
            historical_pe = None
            if pe_data:
                pe_values = [v for v in pe_data.values() if v is not None and v > 0]
                if pe_values:
                    historical_pe = np.mean(pe_values)  # 取平均历史PE
                    if historical_pe <= 10:
                        preliminary_score += 15
                    elif historical_pe <= 15:
                        preliminary_score += 12
                    elif historical_pe <= 20:
                        preliminary_score += 8
                    elif historical_pe <= 30:
                        preliminary_score += 5
            
            return {
                'stock_code': stock_code,
                'stock_name': stock_data['basic_info'].get('stock_name', ''),
                'preliminary_score': preliminary_score,
                'historical_pe': historical_pe,
                'roe_avg': roe_avg * 100 if 'roe_avg' in locals() and roe_avg else None,
                'revenue_growth': avg_growth,
                'debt_ratio': debt_ratio * 100 if debt_ratio else None,
                'current_ratio': current_ratio
            }
            
        except Exception as e:
            logger.error(f"初步筛选股票 {stock_code} 时出错: {e}")
            return {'error': str(e)}
    
    def _assess_potential(self, stock_code: str, preliminary_result: Dict) -> float:
        """
        评估股票潜力分数
        基于多个维度给出潜力评估，用于优先级排序
        """
        potential_score = 0
        
        try:
            # ROE稳定性和趋势 (最高15分)
            stock_data = self.get_stock_metrics(stock_code)
            if not stock_data:
                return 0
                
            metrics = stock_data.get('metrics', {})
            roe_data = metrics.get('roe', {})
            
            if roe_data:
                roe_values = [v for v in roe_data.values() if v is not None]
                
                if len(roe_values) >= 3:
                    # ROE稳定性 (前面几年的ROE波动)
                    roe_std = np.std(roe_values)
                    if roe_std < 2:  # 标准差<2%，非常稳定
                        potential_score += 8
                    elif roe_std < 5:  # 标准差<5%，较稳定
                        potential_score += 5
                    elif roe_std < 8:  # 标准差<8%，一般稳定
                        potential_score += 2
                    
                    # ROE趋势 (是否上升趋势)
                    if len(roe_values) >= 3:
                        recent_roe = np.mean(roe_values[-2:])  # 最近2年平均
                        early_roe = np.mean(roe_values[:2])   # 早期2年平均
                        if recent_roe > early_roe * 1.1:  # 最近比早期高10%+
                            potential_score += 7
                        elif recent_roe > early_roe:  # 上升趋势
                            potential_score += 4
            
            # 营收质量和增长一致性 (最高10分)
            revenue_data = metrics.get('revenue', {})
            if revenue_data:
                revenue_values = [v for v in revenue_data.values() if v is not None and v > 0]
                if len(revenue_values) >= 3:
                    # 营收增长一致性 (连续增长年数)
                    growth_count = 0
                    for i in range(1, len(revenue_values)):
                        if revenue_values[i] > revenue_values[i-1]:
                            growth_count += 1
                    
                    if growth_count == len(revenue_values) - 1:  # 连续增长
                        potential_score += 8
                    elif growth_count >= len(revenue_values) * 0.7:  # 70%年份增长
                        potential_score += 5
                    elif growth_count >= len(revenue_values) * 0.5:  # 50%年份增长
                        potential_score += 2
            
            # 现金流质量 (最高8分)
            cf_data = metrics.get('operating_cash_flow', {})
            if cf_data:
                cf_values = [v for v in cf_data.values() if v is not None]
                if cf_values:
                    positive_cf_ratio = sum(1 for v in cf_values if v > 0) / len(cf_values)
                    if positive_cf_ratio >= 0.8:  # 80%以上年份为正
                        potential_score += 8
                    elif positive_cf_ratio >= 0.6:  # 60%以上年份为正
                        potential_score += 5
                    elif positive_cf_ratio >= 0.4:  # 40%以上年份为正
                        potential_score += 2
            
            # 财务安全边际 (最高7分)
            debt_data = metrics.get('debt_ratio', {})
            current_data = metrics.get('current_ratio', {})
            
            debt_ratio = 1.0  # 默认值
            current_ratio = 0.5  # 默认值
            
            if debt_data:
                debt_values = [v for v in debt_data.values() if v is not None]
                if debt_values:
                    debt_ratio = np.mean(debt_values) / 100.0
            
            if current_data:
                current_values = [v for v in current_data.values() if v is not None]
                if current_values:
                    current_ratio = np.mean(current_values)
            
            # 极低债务率额外加分
            if debt_ratio < 0.2:
                potential_score += 4
            elif debt_ratio < 0.4:
                potential_score += 2
            
            # 优秀流动性额外加分
            if current_ratio > 2.0:
                potential_score += 3
            elif current_ratio > 1.5:
                potential_score += 1
            
            # 行业相对优势 (基于历史PE的合理性，最高5分)
            historical_pe = preliminary_result.get('historical_pe')
            if historical_pe and 5 <= historical_pe <= 25:  # 合理PE范围
                if historical_pe <= 12:
                    potential_score += 5
                elif historical_pe <= 18:
                    potential_score += 3
                elif historical_pe <= 25:
                    potential_score += 1
            
            return min(potential_score, 50)  # 潜力分数上限50分
            
        except Exception as e:
            logger.error(f"评估股票 {stock_code} 潜力时出错: {e}")
            return 0
    
    def comprehensive_evaluation(self, stock_code: str, use_realtime_pe: bool = True) -> Dict:
        """
        综合评估股票
        
        Args:
            stock_code: 股票代码
            use_realtime_pe: 是否使用实时PE（调用Tushare API）
        """
        try:
            stock_data = self.get_stock_metrics(stock_code)
            
            if not stock_data:
                return {'error': f'未找到股票 {stock_code} 的数据'}
            
            pe_api_used = False
            realtime_pe = None
            
            # 根据参数决定是否获取实时PE
            if use_realtime_pe:
                try:
                    realtime_pe = self.get_realtime_pe(stock_code)
                    if realtime_pe is not None:
                        pe_api_used = True
                        # 用实时PE替换数据中的历史PE
                        if 'metrics' not in stock_data:
                            stock_data['metrics'] = {}
                        if 'pe' not in stock_data['metrics']:
                            stock_data['metrics']['pe'] = {}
                        stock_data['metrics']['pe']['current'] = realtime_pe
                        logger.debug(f"使用实时PE: {stock_code} PE={realtime_pe}")
                    else:
                        logger.debug(f"获取实时PE失败，使用历史PE: {stock_code}")
                except Exception as e:
                    logger.warning(f"获取实时PE失败 {stock_code}: {e}，使用历史PE")
            
            # 三位大师的评分
            buffett_result = self.buffett_criteria(stock_data, include_pe_evaluation=True, realtime_pe=realtime_pe)
            munger_result = self.munger_criteria(stock_data, include_pe_evaluation=True, realtime_pe=realtime_pe)
            graham_result = self.graham_criteria(stock_data, include_pe_evaluation=True, realtime_pe=realtime_pe)
            
            # 综合评分
            total_score = (
                buffett_result['score'] * 0.4 +  # 巴菲特权重40%
                munger_result['score'] * 0.3 +   # 芒格权重30%
                graham_result['score'] * 0.3     # 格雷厄姆权重30%
            )
            
            # 投资等级
            if total_score >= 80:
                grade = "A+ 强烈推荐"
            elif total_score >= 70:
                grade = "A 推荐买入"
            elif total_score >= 60:
                grade = "B+ 值得关注"
            elif total_score >= 50:
                grade = "B 谨慎观察"
            else:
                grade = "C 暂不推荐"
            
            evaluation_result = {
                'stock_code': stock_code,
                'stock_name': stock_data['basic_info']['stock_name'],
                'industry': stock_data['basic_info']['industry'],
                'total_score': round(total_score, 1),
                'grade': grade,
                'buffett_analysis': buffett_result,
                'munger_analysis': munger_result,
                'graham_analysis': graham_result,
                'pe_api_used': pe_api_used,  # 标记是否使用了API
                'evaluation_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            # 添加PE信息
            if realtime_pe is not None:
                evaluation_result['realtime_pe'] = realtime_pe
            
            # 添加历史PE信息
            metrics = stock_data.get('metrics', {})
            pe_data = metrics.get('pe', {})
            if pe_data:
                pe_values = [v for v in pe_data.values() if v is not None and v > 0]
                if pe_values:
                    evaluation_result['pe_ratio'] = np.mean(pe_values)
            
            return evaluation_result
            
        except Exception as e:
            logger.error(f"综合评估股票 {stock_code} 时出错: {e}")
            return {'error': str(e)}
    
    def screen_value_stocks(self, min_score: float = 0, limit: int = None, preliminary_threshold: float = 0, test_mode: bool = False, test_count: int = 100) -> List[Dict]:
        """
        智能筛选价值股票 - 两阶段策略提高召回率，降低API调用
        
        Args:
            min_score: 最终最低分数（已废弃，保留兼容性）
            limit: 最终筛选数量限制（已废弃，保留兼容性）
            preliminary_threshold: 初步筛选阈值（已废弃，保留兼容性）
            test_mode: 是否测试模式（只分析少量股票）
            test_count: 测试模式下分析的股票数量
        """
        if test_mode:
            logger.info(f"🧪 测试模式：只分析前 {test_count} 只股票")
        else:
            logger.info(f"🔍 开始智能筛选所有股票（两阶段策略）")
        
        # 获取所有股票代码和名称
        conn = sqlite3.connect(self.db_path)
        stocks = pd.read_sql_query("SELECT stock_code, stock_name FROM stocks", conn)
        conn.close()
        
        if test_mode:
            stocks = stocks.head(test_count)
            logger.info(f"测试模式：限制分析 {len(stocks)} 只股票")
        
        # 第一阶段：基本筛选 + 历史数据初步评分
        logger.info("🔍 第一阶段：基本筛选 + 历史数据初步评分")
        preliminary_candidates = []
        skipped_count = 0
        skip_reasons = {}
        processed = 0
        
        for _, row in stocks.iterrows():
            stock_code = row['stock_code']
            stock_name = row['stock_name']
            
            # 基本筛选
            should_skip, skip_reason = self._should_skip_stock(stock_code, stock_name)
            
            if should_skip:
                skipped_count += 1
                skip_reasons[skip_reason] = skip_reasons.get(skip_reason, 0) + 1
                logger.debug(f"跳过股票: {stock_code} {stock_name} - {skip_reason}")
                continue
            
            # 历史数据初步评分（不使用实时PE）
            try:
                preliminary_result = self.preliminary_screening(stock_code)
                
                if 'error' not in preliminary_result:
                    # 添加潜力评估
                    potential_score = self._assess_potential(stock_code, preliminary_result)
                    preliminary_result['potential_score'] = potential_score
                    preliminary_result['combined_score'] = preliminary_result['preliminary_score'] + potential_score
                    
                    preliminary_candidates.append(preliminary_result)
                    logger.debug(f"初步候选: {stock_code} {preliminary_result['stock_name']} - "
                               f"基础: {preliminary_result['preliminary_score']:.1f}, "
                               f"潜力: {potential_score:.1f}, "
                               f"综合: {preliminary_result['combined_score']:.1f}")
                
                processed += 1
                if processed % 100 == 0:
                    logger.info(f"已初筛 {processed}/{len(stocks)} 只股票，发现 {len(preliminary_candidates)} 只候选")
                    
            except Exception as e:
                logger.error(f"初步评估股票 {stock_code} 时出错: {e}")
                continue
        
        # 按综合得分排序
        preliminary_candidates.sort(key=lambda x: x['combined_score'], reverse=True)
        
        logger.info(f"第一阶段完成:")
        logger.info(f"  - 原始股票: {len(stocks)} 只")
        logger.info(f"  - 跳过股票: {skipped_count} 只")
        logger.info(f"  - 候选股票: {len(preliminary_candidates)} 只")
        if skip_reasons:
            logger.info(f"  - 跳过原因统计:")
            for reason, count in skip_reasons.items():
                logger.info(f"    • {reason}: {count} 只")
        
        if not preliminary_candidates:
            logger.warning("没有找到合适的候选股票")
            return []
        
        # 第二阶段：智能API调用策略
        logger.info("📡 第二阶段：智能实时PE评估")
        
        # 分层策略：优先级越高，越需要准确的实时PE
        high_priority = [c for c in preliminary_candidates if c['combined_score'] >= 65]  # 高潜力股票
        medium_priority = [c for c in preliminary_candidates if 50 <= c['combined_score'] < 65]  # 中等股票
        low_priority = [c for c in preliminary_candidates if c['combined_score'] < 50]  # 低分股票
        
        logger.info(f"候选股票分层:")
        logger.info(f"  - 高优先级(≥65分): {len(high_priority)} 只 (全部使用实时PE)")
        logger.info(f"  - 中优先级(50-64分): {len(medium_priority)} 只 (选择性使用实时PE)")
        logger.info(f"  - 低优先级(<50分): {len(low_priority)} 只 (主要使用历史PE)")
        
        final_results = []
        api_calls = 0
        api_success = 0
        
        # 处理高优先级股票（全部使用实时PE）
        for candidate in high_priority:
            stock_code = candidate['stock_code']
            try:
                final_evaluation = self.comprehensive_evaluation(stock_code, use_realtime_pe=True)
                
                if final_evaluation.get('pe_api_used', False):
                    api_calls += 1
                    api_success += 1
                
                if 'error' not in final_evaluation:
                    final_results.append(final_evaluation)
                    logger.debug(f"✅ 高优先级: {stock_code} - 评分: {final_evaluation['total_score']:.1f}")
                
                # API调用间隔
                if api_calls % 10 == 0:
                    time.sleep(0.5)
                    
            except Exception as e:
                logger.error(f"分析高优先级股票 {stock_code} 时出错: {e}")
                continue
        
        # 处理中优先级股票（选择性使用实时PE）
        medium_with_api = medium_priority[:min(len(medium_priority), 200)]  # 最多200只使用API
        medium_without_api = medium_priority[len(medium_with_api):]
        
        logger.info(f"中优先级处理策略: {len(medium_with_api)} 只使用实时PE, {len(medium_without_api)} 只使用历史PE")
        
        # 有API的中优先级股票
        for candidate in medium_with_api:
            stock_code = candidate['stock_code']
            try:
                final_evaluation = self.comprehensive_evaluation(stock_code, use_realtime_pe=True)
                
                if final_evaluation.get('pe_api_used', False):
                    api_calls += 1
                    api_success += 1
                
                if 'error' not in final_evaluation:
                    final_results.append(final_evaluation)
                
                # API调用间隔
                if api_calls % 20 == 0:
                    time.sleep(1)
                    
            except Exception as e:
                logger.error(f"分析中优先级股票 {stock_code} 时出错: {e}")
                continue
        
        # 无API的中优先级股票
        for candidate in medium_without_api:
            stock_code = candidate['stock_code']
            try:
                final_evaluation = self.comprehensive_evaluation(stock_code, use_realtime_pe=False)
                
                if 'error' not in final_evaluation:
                    final_results.append(final_evaluation)
                    
            except Exception as e:
                logger.error(f"分析中优先级股票 {stock_code} 时出错: {e}")
                continue
        
        # 处理低优先级股票（主要使用历史PE，少量使用实时PE）
        low_with_api = low_priority[:min(len(low_priority), 50)]  # 最多50只使用API
        low_without_api = low_priority[len(low_with_api):]
        
        logger.info(f"低优先级处理策略: {len(low_with_api)} 只使用实时PE, {len(low_without_api)} 只使用历史PE")
        
        # 有API的低优先级股票
        for candidate in low_with_api:
            stock_code = candidate['stock_code']
            try:
                final_evaluation = self.comprehensive_evaluation(stock_code, use_realtime_pe=True)
                
                if final_evaluation.get('pe_api_used', False):
                    api_calls += 1
                    api_success += 1
                
                if 'error' not in final_evaluation:
                    final_results.append(final_evaluation)
                
                # API调用间隔
                if api_calls % 30 == 0:
                    time.sleep(1.5)
                    
            except Exception as e:
                logger.error(f"分析低优先级股票 {stock_code} 时出错: {e}")
                continue
        
        # 无API的低优先级股票
        for candidate in low_without_api:
            stock_code = candidate['stock_code']
            try:
                final_evaluation = self.comprehensive_evaluation(stock_code, use_realtime_pe=False)
                
                if 'error' not in final_evaluation:
                    final_results.append(final_evaluation)
                    
            except Exception as e:
                logger.error(f"分析低优先级股票 {stock_code} 时出错: {e}")
                continue
        
        # 按最终得分排序
        final_results.sort(key=lambda x: x['total_score'], reverse=True)
        
        # 可选：DeepSeek AI分析（只对高分股票）
        if self.deepseek_analyzer and final_results:
            top_for_ai = final_results[:min(50, len(final_results))]
            logger.info(f"🤖 DeepSeek AI分析（分析前{len(top_for_ai)}名）")
            
            for i, result in enumerate(top_for_ai):
                if result['total_score'] < 60:  # 只对60分以上的股票进行AI分析
                    break
                    
                stock_code = result['stock_code']
                try:
                    stock_data = self.get_stock_metrics(stock_code)
                    score_details = {
                        'total_score': result['total_score'],
                        'buffett': result['buffett_analysis'],
                        'munger': result['munger_analysis'],
                        'graham': result['graham_analysis']
                    }
                    
                    ai_analysis = self.deepseek_analyzer.analyze_stock(stock_code, stock_data, score_details)
                    
                    if ai_analysis:
                        result['ai_analysis'] = ai_analysis
                        logger.info(f"✅ AI分析完成: {stock_code}")
                    else:
                        result['ai_analysis'] = "AI分析暂不可用"
                        
                    if i < len(top_for_ai) - 1:
                        time.sleep(2)
                        
                except Exception as e:
                    logger.error(f"AI分析股票 {stock_code} 时出错: {e}")
                    result['ai_analysis'] = f"AI分析出错: {str(e)}"
        
        # 统计信息
        api_success_rate = (api_success / api_calls * 100) if api_calls > 0 else 0
        logger.info(f"🎉 智能筛选完成！")
        logger.info(f"📊 统计信息:")
        logger.info(f"   - 原始股票数: {len(stocks)}")
        logger.info(f"   - 跳过股票数: {skipped_count}")
        logger.info(f"   - 候选股票数: {len(preliminary_candidates)}")
        logger.info(f"   - 成功分析: {len(final_results)} 只")
        logger.info(f"   - API调用: {api_calls} 次，成功率: {api_success_rate:.1f}%")
        logger.info(f"   - API调用节省: {len(preliminary_candidates) - api_calls} 次")
        
        if final_results:
            avg_final_score = np.mean([r['total_score'] for r in final_results])
            pe_used_count = sum(1 for r in final_results if r.get('pe_api_used', False))
            high_score_count = sum(1 for r in final_results if r['total_score'] >= 70)
            medium_score_count = sum(1 for r in final_results if 50 <= r['total_score'] < 70)
            low_score_count = sum(1 for r in final_results if r['total_score'] < 50)
            
            logger.info(f"   - 平均得分: {avg_final_score:.1f}")
            logger.info(f"   - 使用实时PE: {pe_used_count} 只")
            logger.info(f"   - 高分股票(≥70): {high_score_count} 只")
            logger.info(f"   - 中分股票(50-69): {medium_score_count} 只")
            logger.info(f"   - 低分股票(<50): {low_score_count} 只")
        
        return final_results
    
    def generate_report(self, value_stocks: List[Dict], output_file: str = None) -> str:
        """生成投资报告（支持markdown和Excel格式）"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if not output_file:
            base_name = f"value_investment_report_{timestamp}"
            md_file = f"{base_name}.md"
            excel_file = f"{base_name}.xlsx"
        else:
            # 如果用户指定了输出文件，生成对应的Excel文件名
            if output_file.endswith('.md'):
                md_file = output_file
                excel_file = output_file.replace('.md', '.xlsx')
            else:
                md_file = f"{output_file}.md"
                excel_file = f"{output_file}.xlsx"
        
        # 生成Markdown报告
        self._generate_markdown_report(value_stocks, md_file)
        
        # 生成Excel报告
        self._generate_excel_report(value_stocks, excel_file)
        
        logger.info(f"投资报告已保存: {md_file} 和 {excel_file}")
        return md_file
    
    def _generate_markdown_report(self, value_stocks: List[Dict], output_file: str):
        """生成Markdown格式报告"""
        report_lines = [
            "# 🎯 价值投资分析报告",
            f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"**筛选数量**: {len(value_stocks)} 只股票",
            "",
            "## 📊 投资大师评分体系",
            "",
            "### 🏆 巴菲特标准 (权重40%)",
            "- 持续高ROE (>15%)",
            "- 低债务比率 (<30%)",
            "- 稳定盈利能力",
            "- 强劲现金流",
            "",
            "### 🧠 查理芒格标准 (权重30%)",
            "- 优质行业选择",
            "- 高经营效率",
            "- 强定价权（高毛利率）",
            "- 合理估值",
            "",
            "### 📚 格雷厄姆标准 (权重30%)",
            "- 低PE估值 (<12x)",
            "- 低PB比率 (<2x)",
            "- 安全边际充足",
            "- 稳定股息回报",
            "",
            "## 🌟 精选价值股票",
            ""
        ]
        
        for i, stock in enumerate(value_stocks, 1):
            report_lines.extend([
                f"### {i}. {stock['stock_name']} ({stock['stock_code']})",
                f"**综合评分**: {stock['total_score']}/100 - {stock['grade']}",
                f"**所属行业**: {stock['industry']}",
                "",
                "#### 🏆 巴菲特分析",
                f"**得分**: {stock['buffett_analysis']['score']}/100",
                *[f"- {detail}" for detail in stock['buffett_analysis']['details']],
                "",
                "#### 🧠 芒格分析", 
                f"**得分**: {stock['munger_analysis']['score']}/100",
                *[f"- {detail}" for detail in stock['munger_analysis']['details']],
                "",
                "#### 📚 格雷厄姆分析",
                f"**得分**: {stock['graham_analysis']['score']}/100", 
                *[f"- {detail}" for detail in stock['graham_analysis']['details']],
                ""
            ])
            
            # 添加AI分析部分
            if 'ai_analysis' in stock and stock['ai_analysis']:
                if 'AI分析暂不可用' in stock['ai_analysis'] or 'AI分析出错' in stock['ai_analysis']:
                    report_lines.extend([
                        "#### 🤖 AI深度分析",
                        f"**状态**: {stock['ai_analysis']}",
                        ""
                    ])
                else:
                    report_lines.extend([
                        "#### 🤖 AI深度分析",
                        "",
                        stock['ai_analysis'],
                        ""
                    ])
            
            report_lines.extend([
                "---",
                ""
            ])
        
        report_lines.extend([
            "",
            "## ⚠️ 重要声明",
            "",
            "1. 本报告仅基于历史财务数据分析，不构成投资建议",
            "2. 投资有风险，入市需谨慎",
            "3. 建议结合实时市场信息和个人风险承受能力做出投资决策",
            "4. 价值投资需要长期持有，避免短期投机",
            "",
            f"---",
            f"*报告由价值投资Agent生成于 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*"
        ])
        
        report_content = '\n'.join(report_lines)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(report_content)
    
    def _generate_excel_report(self, value_stocks: List[Dict], output_file: str):
        """生成Excel格式报告"""
        try:
            # 创建工作簿
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = Workbook()
            
            # 删除默认工作表
            wb.remove(wb.active)
            
            # 1. 创建概览表
            self._create_overview_sheet(wb, value_stocks)
            
            # 2. 创建详细评分表
            self._create_detailed_scores_sheet(wb, value_stocks)
            
            # 3. 创建财务指标表
            self._create_financial_metrics_sheet(wb, value_stocks)
            
            # 保存文件
            wb.save(output_file)
            logger.info(f"Excel报告已保存到: {output_file}")
            
        except ImportError:
            logger.warning("openpyxl未安装，无法生成Excel报告。请运行: pip install openpyxl")
        except Exception as e:
            logger.error(f"生成Excel报告失败: {e}")
    
    def _create_overview_sheet(self, wb, value_stocks):
        """创建概览工作表"""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        ws = wb.create_sheet("📊 股票概览", 0)
        
        # 设置标题 - 重新设计为更实用的结构
        headers = [
            "排名", "股票代码", "股票名称", "所属行业", 
            "综合评分", "投资等级",
            "ROE(%)", "ROE年份", "债务比率(%)", "债务年份", 
            "流动比率", "流动年份", "毛利率(%)", "毛利年份", 
            "净利率(%)", "净利年份", "市净率", "市净年份",
            "历史PE(年报)", "PE年份", "实时PE(当前)",
            "巴菲特得分", "芒格得分", "格雷厄姆得分", 
            "分析时间"
        ]
        
        # 写入标题行
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据行
        for row, stock in enumerate(value_stocks, 2):
            # 获取详细财务数据
            stock_data = self.get_stock_metrics(stock['stock_code'])
            metrics = stock_data.get('metrics', {}) if stock_data else {}
            
            # 基本信息
            ws.cell(row=row, column=1, value=row-1)  # 排名
            ws.cell(row=row, column=2, value=stock['stock_code'])
            ws.cell(row=row, column=3, value=stock['stock_name'])
            ws.cell(row=row, column=4, value=stock['industry'])
            ws.cell(row=row, column=5, value=round(stock['total_score'], 1))
            ws.cell(row=row, column=6, value=stock['grade'])
            
            # 财务指标（取最新年份数据）
            col_idx = 7
            
            # ROE
            roe_data = metrics.get('roe', {})
            if roe_data:
                latest_year = max(roe_data.keys())
                roe_value = roe_data[latest_year]
                ws.cell(row=row, column=col_idx, value=round(roe_value, 2) if pd.notna(roe_value) else None)
                ws.cell(row=row, column=col_idx+1, value=latest_year)
            col_idx += 2
            
            # 债务比率
            debt_data = metrics.get('debt_ratio', {})
            if debt_data:
                latest_year = max(debt_data.keys())
                debt_value = debt_data[latest_year]
                ws.cell(row=row, column=col_idx, value=round(debt_value, 2) if pd.notna(debt_value) else None)
                ws.cell(row=row, column=col_idx+1, value=latest_year)
            col_idx += 2
            
            # 流动比率
            current_data = metrics.get('current_ratio', {})
            if current_data:
                latest_year = max(current_data.keys())
                current_value = current_data[latest_year]
                ws.cell(row=row, column=col_idx, value=round(current_value, 2) if pd.notna(current_value) else None)
                ws.cell(row=row, column=col_idx+1, value=latest_year)
            col_idx += 2
            
            # 毛利率
            gross_data = metrics.get('gross_margin', {})
            if gross_data:
                latest_year = max(gross_data.keys())
                gross_value = gross_data[latest_year]
                ws.cell(row=row, column=col_idx, value=round(gross_value, 2) if pd.notna(gross_value) else None)
                ws.cell(row=row, column=col_idx+1, value=latest_year)
            col_idx += 2
            
            # 净利率
            net_data = metrics.get('net_margin', {})
            if net_data:
                latest_year = max(net_data.keys())
                net_value = net_data[latest_year]
                ws.cell(row=row, column=col_idx, value=round(net_value, 2) if pd.notna(net_value) else None)
                ws.cell(row=row, column=col_idx+1, value=latest_year)
            col_idx += 2
            
            # 市净率
            pb_data = metrics.get('pb', {})
            if pb_data:
                latest_year = max(pb_data.keys())
                pb_value = pb_data[latest_year]
                ws.cell(row=row, column=col_idx, value=round(pb_value, 2) if pd.notna(pb_value) else None)
                ws.cell(row=row, column=col_idx+1, value=latest_year)
            col_idx += 2
            
            # 历史PE(年报)
            pe_data = metrics.get('pe', {})
            if pe_data:
                latest_year = max(pe_data.keys())
                pe_value = pe_data[latest_year]
                ws.cell(row=row, column=col_idx, value=round(pe_value, 2) if pd.notna(pe_value) else None)
                ws.cell(row=row, column=col_idx+1, value=latest_year)
            col_idx += 2
            
            # 实时PE(当前)
            realtime_pe = stock.get('realtime_pe')
            ws.cell(row=row, column=col_idx, value=round(realtime_pe, 2) if realtime_pe else None)
            col_idx += 1
            
            # 评分信息
            ws.cell(row=row, column=col_idx, value=stock['buffett_analysis']['score'])
            ws.cell(row=row, column=col_idx+1, value=stock['munger_analysis']['score'])
            ws.cell(row=row, column=col_idx+2, value=stock['graham_analysis']['score'])
            ws.cell(row=row, column=col_idx+3, value=stock.get('evaluation_date', ''))
            
            # 根据评分设置行颜色
            if stock['total_score'] >= 80:
                fill_color = "E8F5E8"  # 浅绿色
            elif stock['total_score'] >= 70:
                fill_color = "FFF2CC"  # 浅黄色
            else:
                fill_color = "FFFFFF"  # 白色
                
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_detailed_scores_sheet(self, wb, value_stocks):
        """创建详细评分工作表"""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        ws = wb.create_sheet("📈 详细评分", 1)
        
        # 设置标题
        headers = ["股票代码", "股票名称", "评分项目", "得分详情", "评分说明"]
        
        # 写入标题行
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        current_row = 2
        for stock in value_stocks:
            stock_code = stock['stock_code']
            stock_name = stock['stock_name']
            
            # 巴菲特评分详情
            for detail in stock['buffett_analysis']['details']:
                ws.cell(row=current_row, column=1, value=stock_code)
                ws.cell(row=current_row, column=2, value=stock_name)
                ws.cell(row=current_row, column=3, value="🏆 巴菲特分析")
                ws.cell(row=current_row, column=4, value=detail)
                ws.cell(row=current_row, column=5, value=stock['buffett_analysis']['methodology'])
                current_row += 1
            
            # 芒格评分详情
            for detail in stock['munger_analysis']['details']:
                ws.cell(row=current_row, column=1, value=stock_code)
                ws.cell(row=current_row, column=2, value=stock_name)
                ws.cell(row=current_row, column=3, value="🧠 芒格分析")
                ws.cell(row=current_row, column=4, value=detail)
                ws.cell(row=current_row, column=5, value=stock['munger_analysis']['methodology'])
                current_row += 1
            
            # 格雷厄姆评分详情
            for detail in stock['graham_analysis']['details']:
                ws.cell(row=current_row, column=1, value=stock_code)
                ws.cell(row=current_row, column=2, value=stock_name)
                ws.cell(row=current_row, column=3, value="📚 格雷厄姆分析")
                ws.cell(row=current_row, column=4, value=detail)
                ws.cell(row=current_row, column=5, value=stock['graham_analysis']['methodology'])
                current_row += 1
            
            # 添加分隔行
            current_row += 1
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_financial_metrics_sheet(self, wb, value_stocks):
        """创建财务指标工作表 - 多年数据对比"""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        ws = wb.create_sheet("💰 多年财务指标", 2)
        
        if not value_stocks:
            return
        
        # 构建多年数据表格
        years = [2020, 2021, 2022, 2023, 2024]
        metrics_names = {
            'roe': 'ROE(%)',
            'debt_ratio': '债务比率(%)',
            'current_ratio': '流动比率',
            'gross_margin': '毛利率(%)',
            'net_margin': '净利率(%)',
            'pb': '市净率',
            'pe': 'PE',
            'asset_turnover': '资产周转率',
            'dividend': '股息率(%)'
        }
        
        # 设置标题行
        headers = ["股票代码", "股票名称", "指标"]
        for year in years:
            headers.append(f"{year}年")
        headers.extend(["平均值", "趋势", "评价"])
        
        # 写入标题行
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        current_row = 2
        
        for stock in value_stocks:
            stock_code = stock['stock_code']
            stock_name = stock['stock_name']
            
            # 获取财务数据
            stock_data = self.get_stock_metrics(stock_code)
            metrics = stock_data.get('metrics', {}) if stock_data else {}
            
            for metric_key, metric_name in metrics_names.items():
                if metric_key in metrics:
                    metric_data = metrics[metric_key]
                    
                    # 基本信息
                    ws.cell(row=current_row, column=1, value=stock_code)
                    ws.cell(row=current_row, column=2, value=stock_name)
                    ws.cell(row=current_row, column=3, value=metric_name)
                    
                    # 各年数据
                    values = []
                    col_idx = 4
                    for year in years:
                        value = metric_data.get(year)
                        if value is not None and pd.notna(value):
                            ws.cell(row=current_row, column=col_idx, value=round(value, 2))
                            values.append(value)
                        else:
                            ws.cell(row=current_row, column=col_idx, value=None)
                        col_idx += 1
                    
                    # 计算统计信息
                    if values:
                        avg_value = np.mean(values)
                        trend = self._calculate_trend(values)
                        
                        # 平均值
                        ws.cell(row=current_row, column=col_idx, value=round(avg_value, 2))
                        col_idx += 1
                        
                        # 趋势
                        if trend > 0.1:
                            trend_str = "上升📈"
                        elif trend < -0.1:
                            trend_str = "下降📉"
                        else:
                            trend_str = "稳定➡️"
                        ws.cell(row=current_row, column=col_idx, value=trend_str)
                        col_idx += 1
                        
                        # 评价
                        evaluation = self._evaluate_metric(metric_key, avg_value, trend)
                        ws.cell(row=current_row, column=col_idx, value=evaluation)
                    
                    current_row += 1
            
            # 添加分隔行
            current_row += 1
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 15)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _evaluate_metric(self, metric_key: str, avg_value: float, trend: float) -> str:
        """评价财务指标"""
        evaluations = {
            'roe': {
                'excellent': 20,
                'good': 15,
                'fair': 10,
                'unit': '%'
            },
            'debt_ratio': {
                'excellent': 30,  # 低于30%为优秀
                'good': 50,
                'fair': 70,
                'unit': '%',
                'reverse': True  # 越低越好
            },
            'current_ratio': {
                'excellent': 2.0,
                'good': 1.5,
                'fair': 1.0,
                'unit': ''
            },
            'gross_margin': {
                'excellent': 30,
                'good': 20,
                'fair': 10,
                'unit': '%'
            },
            'net_margin': {
                'excellent': 10,
                'good': 5,
                'fair': 2,
                'unit': '%'
            },
            'pb': {
                'excellent': 2.0,
                'good': 3.0,
                'fair': 5.0,
                'unit': '',
                'reverse': True  # 越低越好
            },
            'pe': {
                'excellent': 15,
                'good': 25,
                'fair': 40,
                'unit': '',
                'reverse': True  # 越低越好
            }
        }
        
        if metric_key not in evaluations:
            return "无评价标准"
        
        standards = evaluations[metric_key]
        is_reverse = standards.get('reverse', False)
        
        if is_reverse:
            # 越低越好的指标
            if avg_value <= standards['excellent']:
                quality = "优秀✨"
            elif avg_value <= standards['good']:
                quality = "良好✅"
            elif avg_value <= standards['fair']:
                quality = "一般📊"
            else:
                quality = "较差⚠️"
        else:
            # 越高越好的指标
            if avg_value >= standards['excellent']:
                quality = "优秀✨"
            elif avg_value >= standards['good']:
                quality = "良好✅"
            elif avg_value >= standards['fair']:
                quality = "一般📊"
            else:
                quality = "较差⚠️"
        
        # 添加趋势信息
        if trend > 0.1:
            trend_desc = " 向好"
        elif trend < -0.1:
            trend_desc = " 走弱"
        else:
            trend_desc = ""
        
        return f"{quality}{trend_desc}"
    
    def _calculate_trend(self, values: List[float]) -> float:
        """计算趋势（正值表示上升，负值表示下降）"""
        if len(values) < 2:
            return 0
        
        x = np.arange(len(values))
        coeffs = np.polyfit(x, values, 1)
        return coeffs[0]  # 返回斜率
    
    def _calculate_growth_rate(self, values: List[float]) -> float:
        """计算增长率"""
        if len(values) < 2:
            return 0
        
        start_value = values[0]
        end_value = values[-1]
        years = len(values) - 1
        
        if start_value <= 0:
            return 0
        
        return (end_value / start_value) ** (1/years) - 1
    
    def analyze_single_stock(self, stock_code: str) -> Dict:
        """分析单个股票"""
        try:
            return self.comprehensive_evaluation(stock_code, use_realtime_pe=True)
        except Exception as e:
            logger.error(f"分析单个股票 {stock_code} 时出错: {e}")
            return {'error': str(e)}
    
    def export_to_excel(self, results: List[Dict], filename: str = None) -> str:
        """导出结果到Excel文件"""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"value_stocks_analysis_{timestamp}.xlsx"
        
        try:
            self._generate_excel_report(results, filename)
            return filename
        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            return ""
    
    def generate_analysis_report(self, results: List[Dict], filename: str = None) -> str:
        """生成分析报告"""
        if not filename:
            filename = "价值投资分析报告.md"
        
        try:
            self._generate_markdown_report(results, filename)
            return filename
        except Exception as e:
            logger.error(f"生成报告失败: {e}")
            return ""

def main():
    # 解析命令行参数
    import argparse
    parser = argparse.ArgumentParser(description='价值投资分析工具')
    parser.add_argument('--test', action='store_true', help='测试模式（只分析少量股票）')
    parser.add_argument('--test-count', type=int, default=100, help='测试模式下分析的股票数量（默认100）')
    parser.add_argument('--output-excel', action='store_true', help='输出Excel文件')
    parser.add_argument('--stock', type=str, help='分析单个股票（股票代码）')
    
    args = parser.parse_args()
    
    try:
        # 初始化分析器
        agent = ValueInvestmentAgent()
        
        if args.stock:
            # 单个股票分析
            logger.info(f"开始分析单个股票: {args.stock}")
            result = agent.analyze_single_stock(args.stock)
            
            if result and 'error' not in result:
                print("\n" + "="*80)
                print(f"股票分析结果: {result['stock_name']} ({result['stock_code']})")
                print("="*80)
                print(f"综合评分: {result['total_score']:.1f}")
                print(f"行业: {result['industry']}")
                if result.get('realtime_pe'):
                    print(f"实时PE: {result['realtime_pe']:.2f}")
                print("="*80)
            else:
                print(f"分析失败: {result.get('error', '未知错误')}")
                
        else:
            # 批量股票筛选
            if args.test:
                logger.info(f"🧪 启动测试模式，分析前{args.test_count}只股票")
            else:
                logger.info("🚀 启动智能价值投资分析（两阶段策略）")
            
            # 执行筛选
            results = agent.screen_value_stocks(
                test_mode=args.test, 
                test_count=args.test_count
            )
            
            if results:
                logger.info(f"✅ 成功分析 {len(results)} 只股票")
                
                # 显示分析摘要
                print("\n" + "="*100)
                print("📊 价值投资分析摘要")
                print("="*100)
                
                # 统计信息
                high_score = [r for r in results if r['total_score'] >= 70]
                medium_score = [r for r in results if 50 <= r['total_score'] < 70]
                
                print(f"总分析股票: {len(results)} 只")
                print(f"高分股票(≥70分): {len(high_score)} 只")
                print(f"中等股票(50-69分): {len(medium_score)} 只")
                print(f"API调用统计: {sum(1 for r in results if r.get('pe_api_used', False))} 次")
                
                # 显示前10名
                print(f"\n🏆 前10名优质股票:")
                print("-"*100)
                print(f"{'排名':<4} {'股票代码':<10} {'股票名称':<20} {'综合评分':<8} {'实时PE':<8} {'行业':<20}")
                print("-"*100)
                
                for i, stock in enumerate(results[:10], 1):
                    realtime_pe = stock.get('realtime_pe', stock.get('pe_ratio', 'N/A'))
                    pe_str = f"{realtime_pe:.1f}" if isinstance(realtime_pe, (int, float)) else str(realtime_pe)
                    print(f"{i:<4} {stock['stock_code']:<10} {stock['stock_name'][:18]:<20} "
                          f"{stock['total_score']:<8.1f} {pe_str:<8} {stock['industry'][:18]:<20}")
                
                # 输出Excel
                if args.output_excel or not args.test:  # 非测试模式默认输出Excel
                    logger.info("📄 开始生成Excel报告...")
                    agent.export_to_excel(results)
                    logger.info("✅ Excel报告生成完成")
                
                # 输出Markdown
                agent.generate_analysis_report(results)
                
                print("\n" + "="*100)
                print("📈 分析完成！")
                if args.output_excel or not args.test:
                    print("📊 Excel报告: value_stocks_analysis.xlsx")
                print("📝 Markdown报告: 价值投资分析报告.md")
                print("="*100)
                
            else:
                logger.warning("⚠️  没有找到符合条件的股票")
                
    except KeyboardInterrupt:
        logger.info("用户中断程序运行")
    except Exception as e:
        logger.error(f"程序运行出错: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main() 