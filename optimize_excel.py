import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from loguru import logger
from datetime import datetime

class ExcelOptimizer:
    def __init__(self, excel_file='stock_analysis_data.xlsx'):
        """初始化Excel优化器"""
        self.excel_file = excel_file
        self.df = None
        self.setup_logger()
        
    def setup_logger(self):
        """设置日志"""
        log_path = "logs"
        os.makedirs(log_path, exist_ok=True)
        logger.add(
            os.path.join(log_path, "excel_optimization_{time}.log"),
            rotation="500 MB",
            encoding="utf-8"
        )
        
    def load_data(self):
        """加载Excel数据"""
        try:
            self.df = pd.read_excel(self.excel_file)
            logger.info(f"成功加载数据：{len(self.df)}行，{len(self.df.columns)}列")
            return True
        except Exception as e:
            logger.error(f"加载Excel文件失败: {e}")
            return False
    
    def create_summary_view(self):
        """创建汇总视图 - 只显示关键信息"""
        if self.df is None:
            return None
            
        # 基本信息列
        basic_cols = ['stock_code', 'stock_name', 'industry', 'need_analysis']
        
        # 计算各指标的最新值和平均值
        summary_data = []
        
        for _, row in self.df.iterrows():
            summary_row = {}
            
            # 基本信息
            for col in basic_cols:
                if col in row:
                    summary_row[col] = row[col]
            
            # ROE汇总
            roe_cols = [col for col in self.df.columns if col.startswith('roe_')]
            roe_values = [row[col] for col in roe_cols if pd.notna(row[col])]
            if roe_values:
                summary_row['roe_最新'] = roe_values[-1]
                summary_row['roe_平均'] = np.mean(roe_values)
                summary_row['roe_趋势'] = '上升' if len(roe_values) > 1 and roe_values[-1] > roe_values[0] else '下降'
            
            # 毛利率汇总
            gm_cols = [col for col in self.df.columns if col.startswith('gross_margin_')]
            gm_values = [row[col] for col in gm_cols if pd.notna(row[col])]
            if gm_values:
                summary_row['毛利率_最新'] = gm_values[-1]
                summary_row['毛利率_平均'] = np.mean(gm_values)
            
            # 净利率汇总
            nm_cols = [col for col in self.df.columns if col.startswith('net_margin_')]
            nm_values = [row[col] for col in nm_cols if pd.notna(row[col])]
            if nm_values:
                summary_row['净利率_最新'] = nm_values[-1]
                summary_row['净利率_平均'] = np.mean(nm_values)
            
            # PE汇总
            pe_cols = [col for col in self.df.columns if col.startswith('pe_')]
            pe_values = [row[col] for col in pe_cols if pd.notna(row[col])]
            if pe_values:
                summary_row['PE_最新'] = pe_values[-1]
                summary_row['PE_平均'] = np.mean(pe_values)
            
            # 股息率汇总
            div_cols = [col for col in self.df.columns if col.startswith('dividend_')]
            div_values = [row[col] for col in div_cols if pd.notna(row[col])]
            if div_values:
                summary_row['股息率_最新'] = div_values[-1]
                summary_row['股息率_平均'] = np.mean(div_values)
            
            # 综合评分（简单评分逻辑）
            score = 0
            if 'roe_平均' in summary_row and summary_row['roe_平均'] > 15:
                score += 20
            if '毛利率_平均' in summary_row and summary_row['毛利率_平均'] > 30:
                score += 20
            if '净利率_平均' in summary_row and summary_row['净利率_平均'] > 10:
                score += 20
            if 'PE_平均' in summary_row and 10 < summary_row['PE_平均'] < 25:
                score += 20
            if '股息率_平均' in summary_row and summary_row['股息率_平均'] > 2:
                score += 20
            
            summary_row['综合评分'] = score
            summary_data.append(summary_row)
        
        return pd.DataFrame(summary_data)
    
    def create_sector_analysis(self):
        """创建行业分析视图"""
        if self.df is None:
            return None
            
        # 按行业分组统计
        sector_stats = []
        
        for industry in self.df['industry'].unique():
            if pd.isna(industry):
                continue
                
            industry_data = self.df[self.df['industry'] == industry]
            
            # 计算行业平均指标
            roe_cols = [col for col in self.df.columns if col.startswith('roe_')]
            pe_cols = [col for col in self.df.columns if col.startswith('pe_')]
            
            sector_row = {
                '行业': industry,
                '公司数量': len(industry_data),
                '平均ROE': industry_data[roe_cols].mean().mean(),
                '平均PE': industry_data[pe_cols].mean().mean(),
                '高ROE公司数': (industry_data[roe_cols].mean(axis=1) > 15).sum(),
                '需要分析数': (industry_data['need_analysis'] == True).sum()
            }
            sector_stats.append(sector_row)
        
        return pd.DataFrame(sector_stats).sort_values('平均ROE', ascending=False)
    
    def create_filtered_views(self):
        """创建筛选视图"""
        if self.df is None:
            return {}
            
        views = {}
        
        # 高ROE股票（ROE均值>15%）
        roe_cols = [col for col in self.df.columns if col.startswith('roe_')]
        high_roe_mask = self.df[roe_cols].mean(axis=1) > 15
        views['高ROE股票'] = self.df[high_roe_mask][['stock_code', 'stock_name', 'industry'] + roe_cols]
        
        # 低PE股票（PE均值<20）
        pe_cols = [col for col in self.df.columns if col.startswith('pe_')]
        low_pe_mask = self.df[pe_cols].mean(axis=1) < 20
        views['低PE股票'] = self.df[low_pe_mask][['stock_code', 'stock_name', 'industry'] + pe_cols]
        
        # 高股息股票（股息率均值>3%）
        div_cols = [col for col in self.df.columns if col.startswith('dividend_')]
        high_div_mask = self.df[div_cols].mean(axis=1) > 3
        views['高股息股票'] = self.df[high_div_mask][['stock_code', 'stock_name', 'industry'] + div_cols]
        
        # 稳定盈利股票（净利率连续正值）
        nm_cols = [col for col in self.df.columns if col.startswith('net_margin_')]
        stable_profit_mask = (self.df[nm_cols] > 0).all(axis=1)
        views['稳定盈利股票'] = self.df[stable_profit_mask][['stock_code', 'stock_name', 'industry'] + nm_cols]
        
        return views
    
    def save_optimized_excel(self, output_file='stock_analysis_optimized.xlsx'):
        """保存优化后的Excel文件"""
        if self.df is None:
            logger.error("没有数据可保存")
            return False
            
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 1. 汇总视图
                summary_df = self.create_summary_view()
                if summary_df is not None:
                    summary_df.to_excel(writer, sheet_name='汇总视图', index=False)
                    logger.info(f"汇总视图已创建，包含{len(summary_df)}行数据")
                
                # 2. 行业分析
                sector_df = self.create_sector_analysis()
                if sector_df is not None:
                    sector_df.to_excel(writer, sheet_name='行业分析', index=False)
                    logger.info(f"行业分析已创建，包含{len(sector_df)}个行业")
                
                # 3. 筛选视图
                filtered_views = self.create_filtered_views()
                for view_name, view_df in filtered_views.items():
                    if not view_df.empty:
                        view_df.to_excel(writer, sheet_name=view_name, index=False)
                        logger.info(f"{view_name}已创建，包含{len(view_df)}只股票")
                
                # 4. 原始数据（可选）
                self.df.to_excel(writer, sheet_name='原始数据', index=False)
                logger.info("原始数据已保留")
            
            # 添加样式
            self._apply_styles(output_file)
            logger.info(f"优化后的Excel文件已保存: {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"保存Excel文件失败: {e}")
            return False
    
    def _apply_styles(self, excel_file):
        """应用Excel样式"""
        try:
            wb = load_workbook(excel_file)
            
            # 为每个工作表添加样式
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 设置标题行样式
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                
                # 设置数据行样式
                for row in range(2, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal="center")
                        
                        # 为数值类型的单元格设置条件格式
                        if isinstance(cell.value, (int, float)):
                            if cell.value is not None and cell.value < 0:
                                cell.font = Font(color="FF0000")  # 负值显示红色
                            elif cell.value is not None and cell.value > 20:
                                cell.font = Font(color="008000")  # 高值显示绿色
                
                # 自动调整列宽
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 20)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_file)
            
        except Exception as e:
            logger.error(f"应用样式失败: {e}")
    
    def generate_analysis_suggestions(self):
        """生成分析建议"""
        if self.df is None:
            return None
            
        suggestions = []
        
        # 1. 高价值股票推荐
        summary_df = self.create_summary_view()
        if summary_df is not None:
            high_score_stocks = summary_df[summary_df['综合评分'] >= 80].sort_values('综合评分', ascending=False)
            if not high_score_stocks.empty:
                suggestions.append("🌟 高价值股票推荐：")
                for _, stock in high_score_stocks.head(10).iterrows():
                    suggestions.append(f"  • {stock['stock_name']}({stock['stock_code']}) - 评分:{stock['综合评分']}")
        
        # 2. 行业机会
        sector_df = self.create_sector_analysis()
        if sector_df is not None:
            top_sectors = sector_df.head(5)
            suggestions.append("\n📈 优势行业：")
            for _, sector in top_sectors.iterrows():
                suggestions.append(f"  • {sector['行业']} - 平均ROE:{sector['平均ROE']:.2f}%")
        
        # 3. 筛选建议
        filtered_views = self.create_filtered_views()
        suggestions.append("\n🔍 筛选建议：")
        for view_name, view_df in filtered_views.items():
            suggestions.append(f"  • {view_name}: {len(view_df)}只股票")
        
        return '\n'.join(suggestions)

def main():
    """主程序"""
    print("📊 A股数据Excel优化工具")
    print("=" * 50)
    
    # 检查原始文件
    if not os.path.exists('stock_analysis_data.xlsx'):
        print("❌ 未找到 stock_analysis_data.xlsx 文件")
        print("请先运行 collect_data.py 收集数据")
        return
    
    # 初始化优化器
    optimizer = ExcelOptimizer()
    
    # 加载数据
    if not optimizer.load_data():
        print("❌ 数据加载失败")
        return
    
    print(f"✅ 成功加载数据：{len(optimizer.df)}只股票，{len(optimizer.df.columns)}列")
    
    # 创建优化文件
    if optimizer.save_optimized_excel():
        print("✅ 优化文件创建成功: stock_analysis_optimized.xlsx")
        
        # 生成分析建议
        suggestions = optimizer.generate_analysis_suggestions()
        if suggestions:
            print("\n📋 分析建议：")
            print(suggestions)
            
            # 保存建议到文件
            with open('analysis_suggestions.txt', 'w', encoding='utf-8') as f:
                f.write(suggestions)
            print("\n💾 分析建议已保存到: analysis_suggestions.txt")
        
        print("\n🎯 优化完成！新文件包含以下工作表：")
        print("  • 汇总视图 - 关键指标汇总")
        print("  • 行业分析 - 行业对比数据")
        print("  • 高ROE股票 - ROE>15%的股票")
        print("  • 低PE股票 - PE<20的股票")  
        print("  • 高股息股票 - 股息率>3%的股票")
        print("  • 稳定盈利股票 - 净利率连续为正的股票")
        print("  • 原始数据 - 完整的原始数据")
        
    else:
        print("❌ 优化文件创建失败")

if __name__ == "__main__":
    main() 